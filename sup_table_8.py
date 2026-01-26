from pathlib import Path
from typing import Dict, Tuple, Union

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

DATA_START_COL = "T8"
REFERENCE_COL = "T6"

THROUGHPUT_CUTOFFS = [5, 10, 20, 30, 40, 50, 60, 70, 80, 90, 100, 200, 300, 400, 500, 1000, 1500]
REF_PANEL_CUTOFFS = [2, 3, 4, 5, 10]


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = out.columns.astype(str).str.strip()
    return out


def _get_data_block(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    df = _normalize_columns(df)
    if DATA_START_COL not in df.columns:
        raise KeyError(f"Missing data start column: {DATA_START_COL}")
    if REFERENCE_COL not in df.columns:
        raise KeyError(f"Missing reference column: {REFERENCE_COL}")
    start_idx = df.columns.get_loc(DATA_START_COL)
    data_df = df.iloc[:, start_idx:]
    return df, data_df


def _track_counts(df: pd.DataFrame) -> pd.Series:
    _, data_df = _get_data_block(df)
    return data_df.notna().sum(axis=0)


def _reference_counts(df: pd.DataFrame) -> Tuple[pd.Series, pd.Series]:
    df, data_df = _get_data_block(df)
    t6 = pd.to_numeric(df[REFERENCE_COL], errors="coerce")
    non_path_mask = t6.isin([1, 2])
    path_mask = t6.isin([4, 5])
    non_path_counts = data_df.loc[non_path_mask].notna().sum(axis=0)
    path_counts = data_df.loc[path_mask].notna().sum(axis=0)
    return non_path_counts, path_counts


def summarize_table(df: pd.DataFrame) -> Dict[str, Union[int, Dict[int, int]]]:
    df = _normalize_columns(df)
    total_counts = _track_counts(df)
    total_tracks = int(len(total_counts))
    tracks_lt_5 = int((total_counts < 5).sum())
    tracks_ge = {cutoff: int((total_counts >= cutoff).sum()) for cutoff in THROUGHPUT_CUTOFFS}

    non_path_counts, path_counts = _reference_counts(df)
    ref_panel_ge = {
        cutoff: int(((non_path_counts >= cutoff) & (path_counts >= cutoff)).sum())
        for cutoff in REF_PANEL_CUTOFFS
    }
    ref_panel_and_total_10 = {
        cutoff: int(
            (
                (total_counts >= 10)
                & (non_path_counts >= cutoff)
                & (path_counts >= cutoff)
            ).sum()
        )
        for cutoff in REF_PANEL_CUTOFFS
    }

    return {
        "total_tracks": total_tracks,
        "tracks_lt_5": tracks_lt_5,
        "tracks_ge": tracks_ge,
        "ref_panel_ge": ref_panel_ge,
        "ref_panel_and_total_10": ref_panel_and_total_10,
    }


def summarize_tables(
    brca1_table: pd.DataFrame, brca2_table: pd.DataFrame
) -> Dict[str, Dict[str, Union[int, Dict[int, int]]]]:
    return {
        "BRCA1": summarize_table(brca1_table),
        "BRCA2": summarize_table(brca2_table),
    }


def write_sup_table_8(
    summary: Dict[str, Dict[str, Union[int, Dict[int, int]]]],
    output_path: str,
) -> None:
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    if output_file.exists():
        wb = load_workbook(output_file)
        if "Sup Table 8" in wb.sheetnames:
            wb.remove(wb["Sup Table 8"])
        ws = wb.create_sheet("Sup Table 8")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sup Table 8"

    ws["A1"] = "Supplementary Table 8: Track throughput"
    ws["B1"] = "BRCA1"
    ws["E1"] = "BRCA2"
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    ws["A2"] = "Functional track throughput"
    ws["B2"] = "N"
    ws["C2"] = "%"
    ws["E2"] = "N"
    ws["F2"] = "%"

    ws["A3"] = "Total number of functional tracks"
    ws["B3"] = summary["BRCA1"]["total_tracks"]
    ws["E3"] = summary["BRCA2"]["total_tracks"]
    ws["C3"] = 100
    ws["F3"] = 100

    def percent(count: int, total: int) -> float:
        return 0.0 if total == 0 else round((count / total) * 100, 2)

    brca1_total_tracks = summary["BRCA1"]["total_tracks"]
    brca2_total_tracks = summary["BRCA2"]["total_tracks"]

    ws["A4"] = "Tracks testing less than 5 variants"
    ws["B4"] = summary["BRCA1"]["tracks_lt_5"]
    ws["E4"] = summary["BRCA2"]["tracks_lt_5"]
    ws["C4"] = percent(summary["BRCA1"]["tracks_lt_5"], brca1_total_tracks)
    ws["F4"] = percent(summary["BRCA2"]["tracks_lt_5"], brca2_total_tracks)

    start_row = 5
    for offset, cutoff in enumerate(THROUGHPUT_CUTOFFS):
        row = start_row + offset
        ws[f"A{row}"] = f"Tracks testing >= {cutoff} variants"
        brca1_val = summary["BRCA1"]["tracks_ge"][cutoff]
        brca2_val = summary["BRCA2"]["tracks_ge"][cutoff]
        ws[f"B{row}"] = brca1_val
        ws[f"E{row}"] = brca2_val
        ws[f"C{row}"] = percent(brca1_val, brca1_total_tracks)
        ws[f"F{row}"] = percent(brca2_val, brca2_total_tracks)

    ws["B23"] = "N"
    ws["C23"] = "%"
    ws["E23"] = "N"
    ws["F23"] = "%"

    for row in (1, 2, 23):
        for col in ("A", "B", "C", "D", "E", "F"):
            cell = ws[f"{col}{row}"]
            if cell.value is not None:
                cell.font = Font(bold=True)

    ref_start_row = 24
    for offset, cutoff in enumerate(REF_PANEL_CUTOFFS):
        row = ref_start_row + offset
        ws[f"A{row}"] = (
            "Tracks with # variants in ENIGMA+ClinVar reference panel "
            f"[non-pathogenic; pathogenic] >= [{cutoff}:{cutoff}]"
        )
        brca1_val = summary["BRCA1"]["ref_panel_ge"][cutoff]
        brca2_val = summary["BRCA2"]["ref_panel_ge"][cutoff]
        ws[f"B{row}"] = brca1_val
        ws[f"E{row}"] = brca2_val
        ws[f"C{row}"] = percent(brca1_val, brca1_total_tracks)
        ws[f"F{row}"] = percent(brca2_val, brca2_total_tracks)

    combo_start_row = 30
    for offset, cutoff in enumerate(REF_PANEL_CUTOFFS):
        row = combo_start_row + offset
        ws[f"A{row}"] = (
            "Tracks meeting the two criteria (#variants tested >= 10 and # variants in "
            "ENIGMA+ClinVar reference panel >= [non-pathogenic; pathogenic] "
            f"[{cutoff}:{cutoff}])"
        )
        brca1_val = summary["BRCA1"]["ref_panel_and_total_10"][cutoff]
        brca2_val = summary["BRCA2"]["ref_panel_and_total_10"][cutoff]
        ws[f"B{row}"] = brca1_val
        ws[f"E{row}"] = brca2_val
        ws[f"C{row}"] = percent(brca1_val, brca1_total_tracks)
        ws[f"F{row}"] = percent(brca2_val, brca2_total_tracks)

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            cell.font = Font(name="Arial", size=10, bold=bool(cell.font.bold))
            if cell.column_letter == "A":
                if cell.row == 1:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is None:
                continue
            if col_letter == "A" and cell.row == 1:
                continue
            max_len = max(max_len, len(str(cell.value)))
        if max_len > 0:
            ws.column_dimensions[col_letter].width = max(12, min(max_len + 2, 120))

    wb.save(output_file)
