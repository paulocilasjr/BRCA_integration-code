#!/usr/bin/env python3
"""
Build Supplementary Tables 18 and 19 from Sup Tables 12/13, EVE, and ACMG inputs.

Adds columns:
  - ACMG BS3/PS3 functional points computation
  - Number of assays
  - ACMG BS3/PS3 Final Functional Points
  - ACMG BS3/PS3 Capped Final Functional Points
  - EVE Score
  - EVE classification
  - ACMG PP3/BP4 in silico predictor points
  - ACMG PM2 points
  - ACMG PP1/BS4 segregation points
  - ACMG BA1/BS1 allele freq points
  - FINAL ACMG POINTS
  - FINAL ClinVar Classification
"""

from __future__ import annotations

import argparse
import ast
import re
from pathlib import Path
from typing import Dict, Iterable, List

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font


FINAL_CODE_COL = "Final functional code label"
EVE_SCORE_COL = "EVE Score"
EVE_CLASS_COL = "EVE classification"
EVE_SOURCE_COL = "EVE Source"
INSILICO_POINTS_COL = "ACMG PP3/BP4 in silico predictor points"
TABLE_TITLES = {
    "Sup Table 18": "Supplementary Table 18: Point system integration for BRCA1 variants",
    "Sup Table 19": "Supplementary Table 19: Point system integration for BRCA2 variants",
}
REFERENCE_NOTE_LP = "Disregard (Reference LP/P)"
REFERENCE_NOTE_LB = "Disregard (Reference LB/B)"
SPECIAL_NOTES: Dict[str, Dict[str, str]] = {
    "BRCA1": {
        "p.R71G": "Disregard this classification: this variant has been classified as LP/P (due to splicing defect)",
        "p.R71K": "Disregard this classification: this variant has been classified as LP/P (due to splicing defect)",
        "p.R71W": "Disregard this classification: this variant has been classified as LP/P (due to splicing defect)",
        "p.R1495K": "Disregard this classification: this variant has been classified as LP/P (due to splicing defect)",
        "p.R1495M": "Disregard this classification: this variant has been classified as LP/P (due to splicing defect)",
        "p.E1559K": "Disregard this classification: this variant has been classified as LP/P (due to splicing defect)",
        "p.E1559Q": "Disregard this classification: this variant has been classified as LP/P (due to splicing defect)",
        "p.A1623G": "Disregard this classification: this variant has been classified as LP/P (due to splicing defect)",
        "p.E1682K": "Potential splicing defect (SpliceAI = 0.57)",
        "p.D1692A": "Disregard this classification: this variant has been classified as LP/P (due to splicing defect)",
        "p.C1787S": "Disregard this classification: this variant has been classified as LP/P (due to co-occurrence with BRCA1 p.Gly1788Asp)",
    },
    "BRCA2": {
        "p.V159M": "Disregard this classification: this variant has been classified as LP/P (due to splicing defect)",
        "p.V211I": "Disregard this classification: this variant has been classified as LP/P (due to splicng defect and co-occurrence with BRCA2 c.7008-2A>T)",
        "p.R2336L": "Disregard this classification: this variant has been classified as LP/P (due to splicing defect)",
        "p.R2336P": "Disregard this classification: this variant has been classified as LP/P (due to splicing defect)",
        "p.R2602T": "Disregard this classification: this variant has been classified as LP/P (due to splicing defect)",
        "p.I2675V": "Disregard this classification: this variant has been classified as LP/P (due to splicing defect)",
        "p.D2679Y": "Disregard this classification: this variant has been classified as LP/P (due to splicing defect)",
    },
}
MANUAL_VALUE_OVERRIDES: Dict[str, Dict[str, Dict[str, object]]] = {
    "BRCA1": {
        "p.R71G": {"ACMG PP1/BS4 segregation points": 8},
    },
    "BRCA2": {
        "p.D2312V": {"ACMG PP1/BS4 segregation points": -4},
        "p.R2336H": {"ACMG PP1/BS4 segregation points": 4},
        "p.Q2829R": {"ACMG PP1/BS4 segregation points": 2},
    },
}
FINAL_CLASS_OVERRIDES: Dict[str, Dict[str, str]] = {
    "BRCA2": {"p.R2336H": "P"},
}


def _norm_col(c: str) -> str:
    return re.sub(r"\s+", " ", str(c).strip().lower())


def _find_col(df: pd.DataFrame, candidates: Iterable[str]) -> str:
    mapping = {_norm_col(c): c for c in df.columns}
    for cand in candidates:
        key = _norm_col(cand)
        if key in mapping:
            return mapping[key]
    for cand in candidates:
        key = _norm_col(cand)
        for k_norm, k_orig in mapping.items():
            if key in k_norm:
                return k_orig
    raise KeyError(f"Could not find any of columns: {list(candidates)}")


def _reference_note_from_t6_and_class(t6_value: object, final_class: object) -> object:
    if pd.isna(t6_value):
        return pd.NA
    tokens = set(re.findall(r"\d+", str(t6_value)))
    final_class_norm = str(final_class).strip().upper()
    if tokens & {"4", "5"} and final_class_norm not in {"LP", "P"}:
        return REFERENCE_NOTE_LP
    if tokens & {"1", "2"} and final_class_norm not in {"LB", "B"}:
        return REFERENCE_NOTE_LB
    return pd.NA


def resolve_sheet_name(xlsx_path: str, preferred: str, fallbacks: List[str]) -> str:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    names = wb.sheetnames
    if preferred in names:
        return preferred
    lower_map = {n.lower(): n for n in names}
    if preferred.lower() in lower_map:
        return lower_map[preferred.lower()]
    for pat in fallbacks:
        pat_l = pat.lower()
        for n in names:
            if pat_l in n.lower():
                return n
    raise ValueError(f"Sheet not found. preferred={preferred!r}. Available: {names}")


def parse_vote_to_score(vote: str) -> int:
    vote = vote.strip()
    if re.match(r"BS3_supporting\b", vote):
        return -1
    if re.match(r"BS3_moderate\b", vote):
        return -2
    if re.match(r"BS3\b", vote):
        return -4
    if re.match(r"PS3_supporting\b", vote):
        return 1
    if re.match(r"PS3_moderate\b", vote):
        return 2
    if re.match(r"PS3\b", vote):
        return 4
    if re.match(r"hypomorph\b", vote, flags=re.IGNORECASE):
        return 0
    return 0


def safe_parse_votes(votes_raw) -> List[str]:
    if pd.isna(votes_raw) or votes_raw == "":
        return []
    try:
        if isinstance(votes_raw, str):
            return ast.literal_eval(votes_raw)
        return votes_raw if isinstance(votes_raw, list) else []
    except Exception:
        return []


def compute_acmg_score(votes: List[str]) -> List[str]:
    scores = []
    for v in votes:
        score = parse_vote_to_score(v)
        match = re.search(r"T\d+", v)
        score_str = f"{score} ({match.group()})" if match else str(score)
        scores.append(score_str)
    return scores


def compute_final_score(score_list: List[str]) -> int:
    try:
        return sum(int(str(s).split()[0]) for s in score_list)
    except Exception:
        return 0


def cap_score(score: int) -> int:
    if score < -4:
        return -4
    if score > 4:
        return 4
    return score


def classify_final_points(score: float) -> str:
    if score >= 10:
        return "P"
    if 6 <= score <= 9:
        return "LP"
    if -1 <= score <= 5:
        return "VUS"
    if -6 <= score <= -2:
        return "LB"
    if score <= -7:
        return "B"
    return ""


def _eve_points(eve_class: object) -> int:
    if pd.isna(eve_class):
        return 0
    eve_class_norm = str(eve_class).strip().lower()
    if eve_class_norm == "benign":
        return -1
    if eve_class_norm == "pathogenic":
        return 1
    return 0


def _parse_eve_variant_columns(df: pd.DataFrame) -> pd.DataFrame:
    if {"T2", "T3", "T4"}.issubset(set(df.columns)):
        return df
    if {"wt_aa", "position", "mt_aa"}.issubset(set(df.columns)):
        df = df.copy()
        df["T2"] = df["wt_aa"]
        df["T3"] = df["position"]
        df["T4"] = df["mt_aa"]
    return df


def _load_eve_predictor(path: str, sheet: str) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=sheet)
    df = _parse_eve_variant_columns(df)
    col_t2 = _find_col(df, ["T2"])
    col_t3 = _find_col(df, ["T3"])
    col_t4 = _find_col(df, ["T4"])
    col_score = _find_col(df, [EVE_SCORE_COL, "EVE_scores_ASM", "EVE score"])
    col_class = _find_col(df, [EVE_CLASS_COL, "EVE_classes_75_pct_retained_ASM", "EVE Class (75% Set)"])

    source_col = None
    try:
        source_col = _find_col(df, [EVE_SOURCE_COL])
    except KeyError:
        pass

    keep_cols = [col_t2, col_t3, col_t4, col_score, col_class]
    if source_col:
        keep_cols.append(source_col)
    out = df[keep_cols].copy()
    rename_cols = {
        col_t2: "T2",
        col_t3: "T3",
        col_t4: "T4",
        col_score: EVE_SCORE_COL,
        col_class: EVE_CLASS_COL,
    }
    if source_col:
        rename_cols[source_col] = EVE_SOURCE_COL
    out = out.rename(columns=rename_cols)
    out = out.dropna(subset=["T2", "T3", "T4"])
    out["T2"] = out["T2"].astype(str).str.strip()
    out["T4"] = out["T4"].astype(str).str.strip()
    out["T3"] = pd.to_numeric(out["T3"], errors="coerce")
    out[EVE_SCORE_COL] = pd.to_numeric(out[EVE_SCORE_COL], errors="coerce")
    out[EVE_CLASS_COL] = out[EVE_CLASS_COL].fillna("No score")
    if EVE_SOURCE_COL not in out.columns:
        out[EVE_SOURCE_COL] = Path(path).name
    out[INSILICO_POINTS_COL] = out[EVE_CLASS_COL].apply(_eve_points)
    return out[["T2", "T3", "T4", EVE_SCORE_COL, EVE_CLASS_COL, INSILICO_POINTS_COL, EVE_SOURCE_COL]]


def _load_acmg_other_points(path: str, sheet: str) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=sheet)
    col_t2 = _find_col(df, ["T2"])
    col_t3 = _find_col(df, ["T3"])
    col_t4 = _find_col(df, ["T4"])
    col_pm2 = _find_col(df, ["ACMG PM2 points"])
    col_pp1 = _find_col(df, ["ACMG PP1/BS4 segregation points"])
    col_ba1 = _find_col(df, ["ACMG BA1/BS1 allele freq points"])
    out = df[[col_t2, col_t3, col_t4, col_pm2, col_pp1, col_ba1]].copy()
    out = out.rename(
        columns={
            col_t2: "T2",
            col_t3: "T3",
            col_t4: "T4",
            col_pm2: "ACMG PM2 points",
            col_pp1: "ACMG PP1/BS4 segregation points",
            col_ba1: "ACMG BA1/BS1 allele freq points",
        }
    )
    out["T2"] = out["T2"].astype(str).str.strip()
    out["T4"] = out["T4"].astype(str).str.strip()
    out["T3"] = pd.to_numeric(out["T3"], errors="coerce")
    return out


def _attach_new_columns(
    base_df: pd.DataFrame,
    predictor_df: pd.DataFrame,
    other_points_df: pd.DataFrame,
    gene_label: str,
) -> pd.DataFrame:
    base = base_df.copy()
    if "Integrated ACMG evidence criteria" in base.columns and FINAL_CODE_COL not in base.columns:
        base = base.rename(columns={"Integrated ACMG evidence criteria": FINAL_CODE_COL})

    df = base.copy()
    df.columns = df.columns.astype(str).str.strip()

    col_votes = next((c for c in df.columns if "all_votes" in _norm_col(c)), None)
    if not col_votes:
        raise ValueError("Could not find an 'All_votes' column in Sup Table 12/13.")
    col_t2 = _find_col(df, ["T2"])
    col_t3 = _find_col(df, ["T3"])
    col_t4 = _find_col(df, ["T4"])

    df["Parsed_votes"] = df[col_votes].apply(safe_parse_votes)
    df["ACMG BS3/PS3 functional points computation"] = df["Parsed_votes"].apply(compute_acmg_score)
    df["Number of assays"] = df["Parsed_votes"].apply(len)
    df["ACMG BS3/PS3 Final Functional Points"] = df["ACMG BS3/PS3 functional points computation"].apply(compute_final_score)
    df["ACMG BS3/PS3 Capped Final Functional Points"] = df["ACMG BS3/PS3 Final Functional Points"].apply(cap_score)

    df = df.drop(columns=["Parsed_votes"])

    # Merge EVE predictor calls.
    df["T2_key"] = df[col_t2].astype(str).str.strip()
    df["T4_key"] = df[col_t4].astype(str).str.strip()
    df["T3_key"] = pd.to_numeric(df[col_t3], errors="coerce")
    df = df.merge(
        predictor_df,
        how="left",
        left_on=["T2_key", "T3_key", "T4_key"],
        right_on=["T2", "T3", "T4"],
        suffixes=("", "_predictor"),
    )

    # Merge ACMG other points
    df = df.merge(
        other_points_df,
        how="left",
        left_on=["T2_key", "T3_key", "T4_key"],
        right_on=["T2", "T3", "T4"],
        suffixes=("", "_other"),
    )

    drop_cols = [
        "T2_key",
        "T3_key",
        "T4_key",
        "T2_predictor",
        "T3_predictor",
        "T4_predictor",
        "T2_other",
        "T3_other",
        "T4_other",
    ]
    for col in drop_cols:
        if col in df.columns:
            df = df.drop(columns=[col])

    col_t5 = _find_col(df, ["T5"])
    col_t6 = _find_col(df, ["T6"])
    for variant, overrides in MANUAL_VALUE_OVERRIDES.get(gene_label, {}).items():
        mask = df[col_t5].astype(str).str.strip().eq(variant)
        if not mask.any():
            continue
        for column_name, value in overrides.items():
            df.loc[mask, column_name] = value

    # Final ACMG points + classification
    points_cols = [
        INSILICO_POINTS_COL,
        "ACMG PM2 points",
        "ACMG PP1/BS4 segregation points",
        "ACMG BA1/BS1 allele freq points",
        "ACMG BS3/PS3 Capped Final Functional Points",
    ]
    for c in points_cols:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    if EVE_CLASS_COL in df.columns:
        df[EVE_CLASS_COL] = df[EVE_CLASS_COL].fillna("No score")

    df["FINAL ACMG POINTS"] = df[points_cols].sum(axis=1)
    df["FINAL ClinVar Classification"] = df["FINAL ACMG POINTS"].apply(classify_final_points)
    for variant, final_class in FINAL_CLASS_OVERRIDES.get(gene_label, {}).items():
        mask = df[col_t5].astype(str).str.strip().eq(variant)
        if mask.any():
            df.loc[mask, "FINAL ClinVar Classification"] = final_class
    df["Notes"] = [
        _reference_note_from_t6_and_class(t6_value, final_class)
        for t6_value, final_class in zip(df[col_t6], df["FINAL ClinVar Classification"])
    ]
    for variant, note in SPECIAL_NOTES.get(gene_label, {}).items():
        mask = df[col_t5].astype(str).str.strip().eq(variant)
        if mask.any():
            df.loc[mask, "Notes"] = note

    # Order: keep original columns, then append new ones in required order
    append_cols = [
        "ACMG BS3/PS3 functional points computation",
        "Number of assays",
        "ACMG BS3/PS3 Final Functional Points",
        "ACMG BS3/PS3 Capped Final Functional Points",
        EVE_SCORE_COL,
        EVE_CLASS_COL,
        INSILICO_POINTS_COL,
        "ACMG PM2 points",
        "ACMG PP1/BS4 segregation points",
        "ACMG BA1/BS1 allele freq points",
        "FINAL ACMG POINTS",
        "FINAL ClinVar Classification",
        "Notes",
    ]
    existing = [c for c in append_cols if c in df.columns]
    base_cols = [c for c in base.columns if c in df.columns]
    df = df[base_cols + existing]
    return df


def _write_df(output_path: str, sheet_name: str, title: str, df: pd.DataFrame) -> None:
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    if output_file.exists():
        wb = load_workbook(output_file)
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        ws = wb.create_sheet(sheet_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

    ws["A1"] = title
    ws["A1"].font = Font(name="Arial", size=10, bold=True)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    def _excel_value(value):
        if isinstance(value, (list, tuple, dict)):
            return str(value)
        try:
            if pd.isna(value):
                return None
        except (TypeError, ValueError):
            pass
        return value

    headers = list(df.columns)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_offset, (_, row) in enumerate(df.iterrows(), start=3):
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_offset, column=col_idx, value=_excel_value(row[header]))
            cell.font = Font(name="Arial", size=10)
            if col_idx in (1, 9, len(headers)):
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb.save(output_file)


def write_sup_tables_18_19(
    sup12_df: pd.DataFrame,
    sup13_df: pd.DataFrame,
    predictor_path: str,
    other_points_path: str,
    output_path: str,
) -> None:
    predictor_b1 = _load_eve_predictor(predictor_path, "BRCA1")
    predictor_b2 = _load_eve_predictor(predictor_path, "BRCA2")
    other_b1 = _load_acmg_other_points(other_points_path, "BRCA1")
    other_b2 = _load_acmg_other_points(other_points_path, "BRCA2")

    out18 = _attach_new_columns(sup12_df, predictor_b1, other_b1, "BRCA1")
    out19 = _attach_new_columns(sup13_df, predictor_b2, other_b2, "BRCA2")

    _write_df(output_path, "Sup Table 18", TABLE_TITLES["Sup Table 18"], out18)
    _write_df(output_path, "Sup Table 19", TABLE_TITLES["Sup Table 19"], out19)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook", help="Input workbook containing Sup Table 12/13")
    ap.add_argument("-o", "--out", default="SupTable18_19.xlsx", help="Output .xlsx")
    ap.add_argument("--sup12", default="Sup Table 12")
    ap.add_argument("--sup13", default="Sup Table 13")
    ap.add_argument(
        "--predictor",
        "--alpha",
        dest="predictor",
        default="dataset/eve/EVE_BRCA12_scores.xlsx",
        help="EVE workbook with BRCA1/BRCA2 tabs. --alpha is retained as a deprecated alias.",
    )
    ap.add_argument(
        "--other-points",
        default="dataset/ACMG_other_points.xlsx",
        help="Workbook with additional ACMG point columns (BRCA1/BRCA2 tabs)",
    )
    args = ap.parse_args()

    sup12_sheet = resolve_sheet_name(args.workbook, args.sup12, ["sup table 12", "table 12"])
    sup13_sheet = resolve_sheet_name(args.workbook, args.sup13, ["sup table 13", "table 13"])

    sup12_df = pd.read_excel(args.workbook, sheet_name=sup12_sheet, header=1)
    sup13_df = pd.read_excel(args.workbook, sheet_name=sup13_sheet, header=1)

    write_sup_tables_18_19(sup12_df, sup13_df, args.predictor, args.other_points, args.out)
    print(f"[OK] Wrote Sup Table 18/19 to {args.out}")


if __name__ == "__main__":
    main()
