from pathlib import Path
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

DATA_START_COL = "T8"
FINAL_CODE_COL = "Final functional code label"

CLASSIFICATION_HIERARCHY = {
    "BS3": 5,
    "BS3_moderate": 4,
    "BS3_supporting": 3,
    "PS3": 5,
    "PS3_moderate": 4,
    "PS3_supporting": 3,
    "Hypomorph": 2,
    "Indeterminate": 1,
}


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = out.columns.astype(str).str.strip()
    return out


def _resolve_track_col(meta_df: pd.DataFrame) -> str:
    meta_df = _normalize_columns(meta_df)
    norm_map = {str(c).strip().lower(): c for c in meta_df.columns}
    candidates = [
        "track #",
        "track#",
        "track number",
        "track no",
        "track id",
        "track",
    ]
    for c in candidates:
        if c in norm_map:
            return norm_map[c]
    return meta_df.columns[0]


def _track_whitelist(meta_df: pd.DataFrame) -> set[str]:
    if meta_df is None or meta_df.empty:
        return set()
    track_col = _resolve_track_col(meta_df)
    whitelist: set[str] = set()
    for val in meta_df[track_col].dropna():
        track_raw = str(val).strip()
        if not track_raw or track_raw.lower() == "nan":
            continue
        if track_raw.startswith("T"):
            whitelist.add(track_raw)
        else:
            whitelist.add(f"T{track_raw}")
    return whitelist


def _filter_data_cols(data_df: pd.DataFrame, meta_df: pd.DataFrame | None) -> pd.DataFrame:
    whitelist = _track_whitelist(meta_df)
    if not whitelist:
        return data_df
    keep = [c for c in data_df.columns if str(c).strip() in whitelist]
    return data_df[keep]


def _get_data_block(df: pd.DataFrame, meta_df: pd.DataFrame | None = None) -> Tuple[pd.DataFrame, pd.DataFrame]:
    df = _normalize_columns(df)
    if DATA_START_COL not in df.columns:
        raise KeyError(f"Missing data start column: {DATA_START_COL}")
    start_idx = df.columns.get_loc(DATA_START_COL)
    data_df = df.iloc[:, start_idx:]
    data_df = _filter_data_cols(data_df, meta_df)
    return df, data_df


def _resolve_track_key(mapping: Dict[str, Dict[str, str]], track: str) -> str:
    if track in mapping:
        return track
    if track.startswith("T") and track[1:] in mapping:
        return track[1:]
    if not track.startswith("T") and f"T{track}" in mapping:
        return f"T{track}"
    return track


def _normalize_class_label(label: str | None) -> str | None:
    if label is None:
        return None
    text = str(label).strip()
    if not text:
        return None
    lower = text.lower()
    if lower in {"indeterminate", "n/a", "na"}:
        return "Indeterminate"
    if lower == "hypomorph":
        return "hypomorph"
    if lower == "bs3_very_strong":
        return "BS3"
    if lower == "ps3_very_strong":
        return "PS3"
    return text


def _is_valid_ratio(var1: int, var2: int, class1: str, class2: str) -> Tuple[int, int, str]:
    if var1 == 0 or var2 == 0:
        return var1, var2, "Indeterminate"
    if var1 / var2 >= 3:
        return var1, var2, class1
    if var2 / var1 >= 3:
        return var2, var1, class2
    return var1, var2, "Indeterminate"


def _get_strongest_classification(array: List[str], category: str | None = None) -> str:
    strongest = "Indeterminate"
    max_strength = 1
    hierarchy = CLASSIFICATION_HIERARCHY
    if category == "PS3":
        hierarchy = {k: v for k, v in CLASSIFICATION_HIERARCHY.items() if k.startswith("PS3") or k == "Indeterminate"}
    if category == "BS3":
        hierarchy = {k: v for k, v in CLASSIFICATION_HIERARCHY.items() if k.startswith("BS3") or k == "Indeterminate"}
    for item in array:
        if item in hierarchy and hierarchy[item] > max_strength:
            max_strength = hierarchy[item]
            strongest = item
    return strongest


def _check_discordance(array: List[str]) -> Tuple[str, str, str, str, str]:
    bs3_count = ps3_count = hypomorph_count = 0
    cleaned = []
    for item in array:
        plain = pd.Series(item).str.replace(r"\(T\d+\)", "", regex=True).iloc[0]
        cleaned.append(plain)
        if "BS3" in plain:
            bs3_count += 1
        if "PS3" in plain:
            ps3_count += 1
        if "hypomorph" in plain.lower():
            hypomorph_count += 1

    bs3_present = bs3_count > 0
    ps3_present = ps3_count > 0
    hypomorph_present = hypomorph_count > 0
    indeterminate_present = any(
        item.strip().lower() == "indeterminate" for item in cleaned
    )
    strongest_call = _get_strongest_classification(cleaned)
    hypo_obs = "Y" if hypomorph_present else "N"

    if bs3_present and ps3_present:
        if bs3_count > ps3_count:
            var1, var2, c1, c2 = bs3_count, ps3_count, "Benign", "Pathogenic"
            strongest = _get_strongest_classification(cleaned, "BS3")
        else:
            var1, var2, c1, c2 = ps3_count, bs3_count, "Pathogenic", "Benign"
            strongest = _get_strongest_classification(cleaned, "PS3")
        _, _, prepon = _is_valid_ratio(var1, var2, c1, c2)
        if prepon == "Indeterminate":
            return "Discordant", prepon, "VUS", "Indeterminate", hypo_obs
        return "Discordant", prepon, strongest, prepon, hypo_obs

    if bs3_present and hypomorph_present:
        if bs3_count > hypomorph_count:
            var1, var2, c1, c2 = bs3_count, hypomorph_count, "Benign", "Hypomorph"
        else:
            var1, var2, c1, c2 = hypomorph_count, bs3_count, "Hypomorph", "Benign"
        _, _, prepon = _is_valid_ratio(var1, var2, c1, c2)
        if prepon == "Indeterminate":
            return "Discordant", prepon, "VUS", "Indeterminate", hypo_obs
        return "Discordant", prepon, strongest_call, prepon, hypo_obs

    if ps3_present and hypomorph_present:
        if ps3_count > hypomorph_count:
            var1, var2, c1, c2 = ps3_count, hypomorph_count, "Pathogenic", "Hypomorph"
        else:
            var1, var2, c1, c2 = hypomorph_count, ps3_count, "Hypomorph", "Pathogenic"
        _, _, prepon = _is_valid_ratio(var1, var2, c1, c2)
        if prepon == "Indeterminate":
            return "Discordant", prepon, "VUS", "Indeterminate", hypo_obs
        return "Discordant", prepon, strongest_call, prepon, hypo_obs

    if bs3_present:
        return "Concordant", "-", strongest_call, "Benign", hypo_obs
    if ps3_present:
        return "Concordant", "-", strongest_call, "Pathogenic", hypo_obs
    if hypomorph_present:
        return "Concordant", "-", "VUS", "Hypomorph", hypo_obs
    if indeterminate_present:
        return "Concordant", "-", "Indeterminate", "Indeterminate", "N"

    return "Indeterminate", "Indeterminate", "VUS", "Indeterminate", "N"


def _build_output(
    brca_table: pd.DataFrame, class_map: Dict[str, Dict[str, str]], meta_df: pd.DataFrame | None = None
) -> pd.DataFrame:
    df, data_df = _get_data_block(brca_table, meta_df)

    out: Dict[int, List[str]] = {}
    for col in data_df.columns:
        for idx, val in data_df[col].items():
            if pd.isna(val):
                continue
            val_num = pd.to_numeric(val, errors="coerce")
            if pd.isna(val_num):
                continue
            key = _resolve_track_key(class_map, str(col))
            cls = None
            if val_num == 1:
                cls = "hypomorph"
            elif val_num == 2:
                cls = class_map.get(key, {}).get("path")
            elif val_num == 0:
                cls = class_map.get(key, {}).get("benign")
            cls = _normalize_class_label(cls)
            if not cls:
                continue
            out.setdefault(idx, []).append(f"{cls}({col})")

    if not out:
        return pd.DataFrame()

    base_cols = [col for col in [f"T{i}" for i in range(1, 8)] if col in df.columns]
    rows = []
    for idx, votes in out.items():
        conc, pre, final_code, notes, hypo = _check_discordance(votes)
        if final_code == "VUS":
            final_code = "Indeterminate"
        base_vals = df.loc[idx, base_cols].to_dict()
        rows.append(
            {
                "INDEX": idx,
                **base_vals,
                "All_votes (track)": str(votes),
                "Concordance": conc,
                "Preponderance of evidence": pre,
                FINAL_CODE_COL: final_code,
                "Functional evidence in favor of": notes,
                "Hypomorph observation": hypo,
            }
        )

    out_df = pd.DataFrame(rows)
    ordered_cols = [
        "INDEX",
        *base_cols,
        "All_votes (track)",
        "Concordance",
        "Preponderance of evidence",
        FINAL_CODE_COL,
        "Functional evidence in favor of",
        "Hypomorph observation",
    ]
    out_df = out_df[ordered_cols]
    if "T3" in out_df.columns and "T4" in out_df.columns:
        out_df["_t3_sort"] = pd.to_numeric(out_df["T3"], errors="coerce")
        out_df["_t4_sort"] = out_df["T4"].astype(str).fillna("")
        out_df = out_df.sort_values(by=["_t3_sort", "_t4_sort"], kind="mergesort")
        out_df = out_df.drop(columns=["_t3_sort", "_t4_sort"])
    return out_df


def _write_sheet(
    df: pd.DataFrame,
    output_path: str,
    sheet_name: str,
    title: str,
) -> None:
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    if output_file.exists():
        wb = load_workbook(output_file)
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        ws = wb.create_sheet(sheet_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

    ws["A1"] = title
    ws["A1"].font = Font(name="Arial", size=10, bold=True)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    if df.empty:
        wb.save(output_file)
        return

    headers = list(df.columns)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)

    start_row = 3
    for r_idx, (_, row) in enumerate(df.iterrows()):
        for c_idx, header in enumerate(headers, start=1):
            ws.cell(row=start_row + r_idx, column=c_idx, value=row[header])

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            cell.font = Font(name="Arial", size=10, bold=bool(cell.font.bold))
            if cell.column_letter == "A":
                if cell.row == 1:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            elif cell.column_letter == "I":
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            elif cell.column_letter in {"J", "K", "L", "M", "N"}:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_letter in ws.column_dimensions:
        ws.column_dimensions[col_letter].width = 18
    ws.column_dimensions["I"].width = 59.14

    for col_letter in ("J", "K", "L", "M", "N"):
        max_len = 0
        for cell in ws[col_letter]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        if max_len > 0:
            ws.column_dimensions[col_letter].width = max(18, min(max_len + 2, 60))

    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 15

    wb.save(output_file)


def write_sup_table_12_13(
    brca1_table: pd.DataFrame,
    brca2_table: pd.DataFrame,
    brca1_class_map: Dict[str, Dict[str, str]],
    brca2_class_map: Dict[str, Dict[str, str]],
    output_path: str,
    brca1_metadata: pd.DataFrame | None = None,
    brca2_metadata: pd.DataFrame | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    df1 = _build_output(brca1_table, brca1_class_map, meta_df=brca1_metadata)
    df2 = _build_output(brca2_table, brca2_class_map, meta_df=brca2_metadata)

    _write_sheet(
        df1,
        output_path,
        "Sup Table 12",
        "Supplementary Table 12: Assignment of functional evidence code for BRCA1 variants",
    )
    _write_sheet(
        df2,
        output_path,
        "Sup Table 13",
        "Supplementary Table 13: Assignment of functional evidence code for BRCA2 variants",
    )

    return df1, df2
