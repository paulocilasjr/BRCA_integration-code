"""Validation helpers for publication inputs and generated workbooks."""

from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


SOURCE_SHEETS = tuple(f"Sup Table {number}" for number in range(1, 7))
EVE_SHEETS = ("README", "BRCA1", "BRCA2")
OTHER_POINTS_SHEETS = ("BRCA1", "BRCA2")
GENERATED_SHEETS = (
    "Sup Table 7",
    "Sup Table 8",
    "Sup Table 9",
    "Sup Table 10",
    "Sup Table 11",
    "Sup Table 12",
    "Sup Table 13",
    "Supp Table 14",
    "Supp Table 15",
    "SuppTable 16",
    "Supp Table 17",
    "Sup Table 18",
    "Sup Table 19",
)

# These row counts include title/header rows and are invariant for the deposited
# source workbooks.  They catch truncated or partially written primary tables.
PRIMARY_OUTPUT_ROWS = {
    "Sup Table 7": 188,
    "Sup Table 8": 150,
    "Sup Table 12": 3_249,
    "Sup Table 13": 6_179,
    "Sup Table 18": 3_249,
    "Sup Table 19": 6_179,
}


class ValidationError(RuntimeError):
    """Raised when a required artifact is absent, malformed, or incomplete."""


def _require_workbook(path: Path, required_sheets: Iterable[str], label: str) -> None:
    if not path.is_file():
        raise ValidationError(f"Missing {label}: {path}")
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
    except Exception as exc:
        raise ValidationError(f"Could not open {label} {path}: {exc}") from exc
    try:
        missing = [sheet for sheet in required_sheets if sheet not in workbook.sheetnames]
        if missing:
            raise ValidationError(f"{label} {path} is missing sheets: {missing}")
    finally:
        workbook.close()


def validate_inputs(
    source_workbook: str | Path,
    eve_workbook: str | Path,
    other_points_workbook: str | Path,
) -> None:
    """Fail early when any required input workbook is absent or malformed."""

    _require_workbook(Path(source_workbook), SOURCE_SHEETS, "source workbook")
    _require_workbook(Path(eve_workbook), EVE_SHEETS, "EVE workbook")
    _require_workbook(
        Path(other_points_workbook), OTHER_POINTS_SHEETS, "ACMG points workbook"
    )


def validate_generated_workbook(
    path: str | Path, *, enforce_publication_rows: bool = True
) -> None:
    """Check that a generated workbook is complete and uses the EVE columns."""

    path = Path(path)
    _require_workbook(path, GENERATED_SHEETS, "generated workbook")
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        if enforce_publication_rows:
            for sheet_name, expected_rows in PRIMARY_OUTPUT_ROWS.items():
                observed = workbook[sheet_name].max_row
                if observed != expected_rows:
                    raise ValidationError(
                        f"{sheet_name} has {observed:,} rows; expected {expected_rows:,}."
                    )

        for sheet_name in ("Sup Table 18", "Sup Table 19"):
            worksheet = workbook[sheet_name]
            headers = {cell.value for cell in worksheet[2] if cell.value is not None}
            required = {"EVE Score", "EVE classification", "FINAL ClinVar Classification"}
            missing = sorted(required - headers)
            if missing:
                raise ValidationError(f"{sheet_name} is missing columns: {missing}")
            deprecated = {"Alpha Missense Score", "Alpha missense classification"} & headers
            if deprecated:
                raise ValidationError(
                    f"{sheet_name} contains deprecated AlphaMissense columns: "
                    f"{sorted(deprecated)}"
                )

        error_tokens = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?"}
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.data_type == "e" or cell.value in error_tokens:
                        raise ValidationError(
                            f"Spreadsheet error in {worksheet.title}!{cell.coordinate}: "
                            f"{cell.value}"
                        )
    finally:
        workbook.close()
