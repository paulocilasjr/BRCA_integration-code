from pathlib import Path
from typing import Dict, List, Tuple

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from sklearn.model_selection import StratifiedKFold

DATA_START_COL = "T8"
REFERENCE_COL = "T6"
EXCLUDED_T6_VARIANTS = {"p.M1I", "p.M1K", "p.M1R", "p.M1T", "p.M1V"}


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = out.columns.astype(str).str.strip()
    return out


def _resolve_track_col(meta_df: pd.DataFrame) -> str:
    meta_df = _normalize_columns(meta_df)
    norm_map = {str(c).strip().lower(): c for c in meta_df.columns}
    candidates = [
        "track #",
        "track#",
        "track number",
        "track no",
        "track id",
        "track",
    ]
    for c in candidates:
        if c in norm_map:
            return norm_map[c]
    return meta_df.columns[0]


def _track_whitelist(meta_df: pd.DataFrame) -> set[str]:
    if meta_df is None or meta_df.empty:
        return set()
    track_col = _resolve_track_col(meta_df)
    whitelist: set[str] = set()
    for val in meta_df[track_col].dropna():
        track_raw = str(val).strip()
        if not track_raw or track_raw.lower() == "nan":
            continue
        if track_raw.startswith("T"):
            whitelist.add(track_raw)
        else:
            whitelist.add(f"T{track_raw}")
    return whitelist


def _filter_data_cols(data_df: pd.DataFrame, meta_df: pd.DataFrame | None) -> pd.DataFrame:
    whitelist = _track_whitelist(meta_df)
    if not whitelist:
        return data_df
    keep = [c for c in data_df.columns if str(c).strip() in whitelist]
    return data_df[keep]


def _norm_t6(series: pd.Series) -> pd.Series:
    def _norm(val: object) -> str:
        if pd.isna(val):
            return ""
        s = str(val).strip().replace(" ", "")
        if s.endswith(".0"):
            s = s[:-2]
        return s
    return series.map(_norm)


def _t6_masks(series: pd.Series, t5_series: pd.Series | None = None) -> Tuple[pd.Series, pd.Series]:
    clean = _norm_t6(series)
    if t5_series is not None:
        mask = t5_series.astype(str).str.strip().isin(EXCLUDED_T6_VARIANTS)
        if mask.any():
            clean = clean.mask(mask, "")
    path_mask = clean.isin({"4", "5", "4;5"})
    ben_mask = clean.isin({"1", "2", "1;2"})
    return path_mask, ben_mask


def _get_data_block(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    df = _normalize_columns(df)
    if DATA_START_COL not in df.columns:
        raise KeyError(f"Missing data start column: {DATA_START_COL}")
    if REFERENCE_COL not in df.columns:
        raise KeyError(f"Missing reference column: {REFERENCE_COL}")
    start_idx = df.columns.get_loc(DATA_START_COL)
    data_df = df.iloc[:, start_idx:]
    return df, data_df


def _wilson_score_interval(p: float, n: int, z: float = 1.96) -> Tuple[float, float]:
    if n == 0:
        return 0.0, 0.0
    denom = 1 + z**2 / n
    center = (p + z**2 / (2 * n)) / denom
    margin = z * ((p * (1 - p) / n) + (z**2 / (4 * n**2))) ** 0.5 / denom
    return max(0.0, center - margin), min(1.0, center + margin)


def _classify_odds(odds: float) -> str:
    if 0.0001 < odds < 0.0029:
        return "BS3_very_strong"
    if odds < 0.053:
        return "BS3"
    if odds < 0.23:
        return "BS3_moderate"
    if odds < 0.48:
        return "BS3_supporting"
    if odds > 350:
        return "PS3_very_strong"
    if odds > 18.7:
        return "PS3"
    if odds > 4.3:
        return "PS3_moderate"
    if odds > 2.1:
        return "PS3_supporting"
    return "indeterminate"


def _get_final_classification(classification_str: str) -> float:
    if pd.isna(classification_str) or str(classification_str).strip() == "":
        return np.nan
    classifications = [item.split(" (")[0] for item in classification_str.split("; ")]
    if all(c.startswith("BS3") for c in classifications):
        return 0
    if all(c.startswith("PS3") for c in classifications):
        return 2
    if all(c in ["indeterminate", "hypomorph"] for c in classifications):
        return 1
    bs3_count = sum(1 for c in classifications if c.startswith("BS3"))
    ps3_count = sum(1 for c in classifications if c.startswith("PS3"))
    if ps3_count >= 3 * bs3_count:
        return 2
    if bs3_count >= 3 * ps3_count:
        return 0
    return 1


def _prepare_df(df: pd.DataFrame, meta_df: pd.DataFrame | None = None) -> pd.DataFrame:
    df = _normalize_columns(df)
    if DATA_START_COL not in df.columns:
        raise KeyError(f"Missing data start column: {DATA_START_COL}")
    if REFERENCE_COL not in df.columns:
        raise KeyError(f"Missing reference column: {REFERENCE_COL}")
    start_idx = df.columns.get_loc(DATA_START_COL)
    df = df[df[REFERENCE_COL].notna()].copy()
    data_df = df.iloc[:, start_idx:]
    data_df = _filter_data_cols(data_df, meta_df)
    df = df[data_df.notna().any(axis=1)].copy()
    return df


def _mcc(tp: int, fp: int, tn: int, fn: int) -> float:
    denom = (tp + fp) * (tp + fn) * (tn + fp) * (tn + fn)
    if denom == 0:
        return 0.0
    return (tp * tn - fp * fn) / np.sqrt(denom)


def _reference_labels(df: pd.DataFrame) -> pd.Series:
    path_mask, ben_mask = _t6_masks(df[REFERENCE_COL], df.get("T5"))
    labels = pd.Series(pd.NA, index=df.index, dtype="object")
    labels.loc[path_mask] = "pathogenic"
    labels.loc[ben_mask] = "benign"
    return labels


def _label_counts(df: pd.DataFrame) -> Tuple[int, int]:
    labels = _reference_labels(df)
    return int((labels == "pathogenic").sum()), int((labels == "benign").sum())


def _compute_fold_metrics(df: pd.DataFrame, n_splits: int, meta_df: pd.DataFrame | None = None) -> List[Dict[str, float]]:
    df = _prepare_df(df, meta_df)
    labels = _reference_labels(df)
    df = df.loc[labels.notna()].copy()
    labels = labels.loc[df.index]
    start_col_index = df.columns.get_loc(DATA_START_COL)
    assay_cols = list(df.columns[start_col_index:])
    assay_cols = list(_filter_data_cols(df[assay_cols], meta_df).columns)

    kf = StratifiedKFold(n_splits=n_splits, shuffle=True, random_state=42)
    metrics_list: List[Dict[str, float]] = []

    for fold, (train_index, test_index) in enumerate(kf.split(df, labels), 1):
        train_df = df.iloc[train_index].copy()
        test_df = df.iloc[test_index].copy()
        train_pathogenic, train_benign = _label_counts(train_df)
        test_pathogenic, test_benign = _label_counts(test_df)

        specificity_sensitivity = {}
        for column in assay_cols:
            path_mask, ben_mask = _t6_masks(train_df[REFERENCE_COL], train_df.get("T5"))
            tp = ((train_df[column] == 2) & path_mask).sum()
            tn = ((train_df[column] == 0) & ben_mask).sum()
            fp = ((train_df[column] == 2) & ~ben_mask).sum()
            fn = ((train_df[column] == 0) & ~path_mask).sum()

            p1_denom = tp + tn + fp + fn
            p1 = (tp + fn) / p1_denom if p1_denom > 0 else 0.0
            p2_path = (tp + fp) / (tp + fp + 0.5) if (tp + fp + 0.5) > 0 else 0.0
            p2_benign = 0.5 / (tn + fn + 0.5) if (tn + fn + 0.5) > 0 else 0.0
            oddspath_path = (p2_path * (1 - p1)) / ((1 - p2_path) * p1) if (1 - p2_path) * p1 != 0 else 0.0
            oddspath_benign = (p2_benign * (1 - p1)) / ((1 - p2_benign) * p1) if (1 - p2_benign) * p1 != 0 else 0.0

            specificity_sensitivity[column] = {
                "acmg_benign": _classify_odds(oddspath_benign),
                "acmg_path": _classify_odds(oddspath_path),
            }

        def classify_variant(row: pd.Series) -> str:
            classification = []
            for col in assay_cols:
                if row[col] == 2:
                    classification.append(f"{specificity_sensitivity[col]['acmg_path']} ({col})")
                elif row[col] == 1:
                    classification.append(f"hypomorph ({col})")
                elif row[col] == 0:
                    classification.append(f"{specificity_sensitivity[col]['acmg_benign']} ({col})")
            return "; ".join(classification) if classification else ""

        classification_str = test_df.apply(classify_variant, axis=1)
        final_class = classification_str.apply(_get_final_classification)
        valid = classification_str.ne("") & classification_str.notna()

        t6 = test_df.loc[valid, REFERENCE_COL]
        final_valid = final_class.loc[valid]

        pathogenic, benign = _t6_masks(t6, test_df.loc[valid].get("T5"))

        tp = int(((final_valid == 2) & pathogenic).sum())
        tn = int(((final_valid == 0) & benign).sum())
        fp = int(((final_valid == 2) & benign).sum())
        fn = int(((final_valid == 0) & pathogenic).sum())
        no_call = int((final_valid == 1).sum())
        total = int(tp + tn + fp + fn + no_call)
        no_call_rate = no_call / total if total > 0 else 0.0

        sens_denom = tp + fn
        sensitivity = tp / sens_denom if sens_denom > 0 else 0.0
        sens_low, sens_high = _wilson_score_interval(sensitivity, sens_denom)

        spec_denom = tn + fp
        specificity = tn / spec_denom if spec_denom > 0 else 0.0
        spec_low, spec_high = _wilson_score_interval(specificity, spec_denom)

        metrics_list.append(
            {
                "fold": fold,
                "train_pathogenic": train_pathogenic,
                "train_benign": train_benign,
                "test_pathogenic": test_pathogenic,
                "test_benign": test_benign,
                "sensitivity": round(sensitivity, 2),
                "sensitivity_lower_95CI": round(sens_low, 2),
                "sensitivity_upper_95CI": round(sens_high, 2),
                "specificity": round(specificity, 2),
                "specificity_lower_95CI": round(spec_low, 2),
                "specificity_upper_95CI": round(spec_high, 2),
                "TP": tp,
                "FP": fp,
                "TN": tn,
                "FN": fn,
                "No Call": no_call,
                "No Call Rate": round(no_call_rate, 2),
                "Total": total,
                "MCC": round(_mcc(tp, fp, tn, fn), 2),
            }
        )

    return metrics_list


def _average_row(metrics: List[Dict[str, float]]) -> Dict[str, float]:
    keys = [k for k in metrics[0].keys() if k != "fold"]
    avg: Dict[str, float] = {"fold": "AVG"}
    no_avg_keys = {
        "train_pathogenic",
        "train_benign",
        "test_pathogenic",
        "test_benign",
        "TP",
        "FP",
        "TN",
        "FN",
        "No Call",
        "Total",
    }
    for key in keys:
        if key in no_avg_keys:
            avg[key] = ""
            continue
        values = [row[key] for row in metrics]
        avg_val = float(np.mean(values))
        avg[key] = round(avg_val, 2) if isinstance(avg_val, float) else avg_val
    return avg


def _write_block(
    ws,
    start_row: int,
    title: str,
    k_label: str,
    metrics: List[Dict[str, float]],
    include_header: bool = True,
) -> None:
    ws[f"A{start_row}"] = title
    ws[f"B{start_row}"] = k_label
    ws[f"A{start_row}"].font = Font(bold=True)
    ws[f"B{start_row}"].font = Font(bold=True)

    headers = [
        "fold",
        "train_pathogenic",
        "train_benign",
        "test_pathogenic",
        "test_benign",
        "sensitivity",
        "sensitivity_lower_95CI",
        "sensitivity_upper_95CI",
        "specificity",
        "specificity_lower_95CI",
        "specificity_upper_95CI",
        "TP",
        "FP",
        "TN",
        "FN",
        "No Call",
        "No Call Rate",
        "Total",
        "MCC",
    ]

    if include_header:
        header_row = start_row + 1
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
        data_start = header_row + 1
    else:
        data_start = start_row + 2

    for idx, row in enumerate(metrics):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=data_start + idx, column=col_idx, value=row[header])

    avg_row = _average_row(metrics)
    avg_row_index = data_start + len(metrics)
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=avg_row_index, column=col_idx, value=avg_row[header])


def write_sup_table_11(
    brca1_table: pd.DataFrame,
    brca2_table: pd.DataFrame,
    output_path: str,
    brca1_metadata: pd.DataFrame | None = None,
    brca2_metadata: pd.DataFrame | None = None,
) -> None:
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    if output_file.exists():
        wb = load_workbook(output_file)
        if "Sup Table 11" in wb.sheetnames:
            wb.remove(wb["Sup Table 11"])
        ws = wb.create_sheet("Sup Table 11")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sup Table 11"

    ws["A1"] = "Supplementary Table 11: stratified k fold cross-validation"
    ws["A1"].font = Font(bold=True)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    metrics_brca1 = _compute_fold_metrics(brca1_table, n_splits=10, meta_df=brca1_metadata)
    _write_block(ws, 2, "BRCA1", "K=10", metrics_brca1)

    metrics_brca2 = _compute_fold_metrics(brca2_table, n_splits=5, meta_df=brca2_metadata)
    _write_block(ws, 16, "BRCA2", "K=5", metrics_brca2, include_header=False)

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            cell.font = Font(name="Arial", size=10, bold=bool(cell.font.bold))
            if cell.column_letter == "A":
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_letter in (
        "A",
        "B",
        "C",
        "D",
        "E",
        "F",
        "G",
        "H",
        "I",
        "J",
        "K",
        "L",
        "M",
        "N",
        "O",
        "P",
        "Q",
        "R",
        "S",
        "T",
    ):
        ws.column_dimensions[col_letter].width = 18

    for row in (14, 23):
        for col in (
            "A",
            "B",
            "C",
            "D",
            "E",
            "F",
            "G",
            "H",
            "I",
            "J",
            "K",
            "L",
            "M",
            "N",
            "O",
            "P",
            "Q",
            "R",
            "S",
            "T",
        ):
            cell = ws[f"{col}{row}"]
            if cell.value is not None:
                cell.font = Font(name="Arial", size=10, bold=True)

    wb.save(output_file)
