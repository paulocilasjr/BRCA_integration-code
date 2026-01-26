#!/usr/bin/env python3
"""
Build Supplementary Table 14 (Frequency + Enrichment ratios by substitution TYPE)
from:
  - Sup Table 1: BRCA1_table (variant table with T1..T7 metadata + T8+ assay tracks)
  - Sup Table 3: BRCA1_metadata (track metadata; used here for consistency checks)
  - Sup Table 5: BRCA1_Reference_panel (reference variants; used to exclude)
  - Sup Table 12: BRCA1 functional evidence assignment (used to stratify VUS into:
        Functional Impact, Hypomorph, Normal function)

Output:
  - An .xlsx with a single sheet "Sup Table 14" formatted like the provided screenshot.

Usage:
  python build_sup_table14.py input_workbook.xlsx -o SupTable14.xlsx

Optional sheet overrides:
  --sup1 "Sup Table 1" --sup3 "Sup Table 3" --sup5 "Sup Table 5" --sup12 "Sup Table 12"
"""

import argparse
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Set, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---- Sup Table 14/15 layout builder (module API) ----
PANEL1_FILL = None
PANEL2_FILL = "F3D7CF"
PANEL3_FILL = "F7E7B7"
PANEL4_FILL = "CFE3F5"

PANEL2_RED = {"WG", "CW", "VD", "WS", "VE", "CF", "AP", "GW", "HP", "IN"}
PANEL3_RED = {"PQ", "YH", "VE", "AD", "MT", "SI"}
PANEL3_BLUE = {"LI", "LR"}
PANEL4_RED = {"TM", "LS", "MI", "SF", "ND", "GS", "KR", "LW"}
DATA_START_ROW = 4


def _norm_col(c: str) -> str:
    return re.sub(r"\s+", " ", str(c).strip().lower())


def _find_col(df: pd.DataFrame, candidates: Iterable[str]) -> str:
    mapping = {_norm_col(c): c for c in df.columns}
    for cand in candidates:
        key = _norm_col(cand)
        if key in mapping:
            return mapping[key]
    for cand in candidates:
        key = _norm_col(cand)
        for k_norm, k_orig in mapping.items():
            if key in k_norm:
                return k_orig
    raise KeyError(f"Could not find any of columns: {list(candidates)}")


def _extract_code(var: str) -> Optional[str]:
    return extract_type(var)


def _counts_freq(vus_df: pd.DataFrame, variant_set: set[str]) -> Tuple[pd.Series, pd.Series]:
    sub = vus_df[vus_df["T5"].isin(variant_set)].copy()
    sub["TYPE"] = sub["T5"].map(_extract_code)
    counts = sub["TYPE"].value_counts().sort_index()
    freqs = (counts / len(sub)).round(3) if len(sub) > 0 else counts * 0
    return counts, freqs


def _panel_data_from_sup12(
    brca_table: pd.DataFrame, sup_df: pd.DataFrame, ref_variants: set[str]
) -> Dict[str, pd.DataFrame | Dict[str, set[str]]]:
    brca_table = brca_table.copy()
    brca_table.columns = [str(c).strip() for c in brca_table.columns]
    sup_df = sup_df.copy()
    sup_df.columns = [str(c).strip() for c in sup_df.columns]

    col_t5 = _find_col(sup_df, ["T5"])
    col_fe = _find_col(sup_df, ["Functional evidence in favor of"])
    col_hypo = _find_col(sup_df, ["Hypomorph observation"])
    col_criteria = None
    try:
        col_criteria = _find_col(sup_df, ["Integrated ACMG evidence criteria"])
    except KeyError:
        col_criteria = None

    sup_df = sup_df[[c for c in [col_t5, col_fe, col_hypo, col_criteria] if c]].dropna(subset=[col_t5]).copy()
    sup_df[col_t5] = sup_df[col_t5].astype(str)
    sup_df[col_fe] = sup_df[col_fe].astype(str).str.strip()
    sup_df[col_hypo] = sup_df[col_hypo].astype(str).str.strip().str.upper()
    if col_criteria:
        sup_df[col_criteria] = sup_df[col_criteria].astype(str).str.strip()

    # Build VUS universe from Sup1 (T6 NaN) excluding Sup5 ref variants
    col_t5_1 = _find_col(brca_table, ["T5"])
    col_t6_1 = _find_col(brca_table, ["T6"])
    sup1_min = brca_table[[col_t5_1, col_t6_1]].dropna(subset=[col_t5_1]).copy()
    sup1_min[col_t5_1] = sup1_min[col_t5_1].astype(str)
    vus_sup1 = sup1_min[sup1_min[col_t6_1].isna()].copy()
    vus_sup1 = vus_sup1[~vus_sup1[col_t5_1].isin(ref_variants)]
    vus_variants = set(vus_sup1[col_t5_1].tolist())

    sup_df_vus = sup_df[sup_df[col_t5].isin(vus_variants)].copy()

    hypo_vars = set(
        sup_df_vus.loc[
            sup_df_vus[col_fe].str.contains("hypomorph", case=False, na=False),
            col_t5,
        ]
    )
    func_vars = set(
        sup_df_vus.loc[
            sup_df_vus[col_fe].str.contains("pathogenic", case=False, na=False)
            | (sup_df_vus[col_criteria].str.contains("PS3", case=False, na=False) if col_criteria else False),
            col_t5,
        ]
    ) - hypo_vars
    norm_vars = set(
        sup_df_vus.loc[
            sup_df_vus[col_fe].str.contains("benign", case=False, na=False)
            | (sup_df_vus[col_criteria].str.contains("BS3", case=False, na=False) if col_criteria else False),
            col_t5,
        ]
    ) - hypo_vars

    n_all = len(sup_df_vus)
    n_func = len(func_vars)
    n_hypo = len(hypo_vars)
    n_norm = len(norm_vars)

    counts_all, freqs_all = _counts_freq(sup_df_vus, set(sup_df_vus[col_t5]))
    counts_func, freqs_func = _counts_freq(sup_df_vus, func_vars)
    counts_hypo, freqs_hypo = _counts_freq(sup_df_vus, hypo_vars)
    counts_norm, freqs_norm = _counts_freq(sup_df_vus, norm_vars)

    types = sorted(counts_all.index.tolist())
    ref_freq = freqs_all.to_dict()

    def make_df(counts: pd.Series, freqs: pd.Series) -> pd.DataFrame:
        df = pd.DataFrame(index=types)
        df["TYPE"] = df.index
        df["c"] = counts.reindex(types, fill_value=0).astype(int)
        df["f"] = freqs.reindex(types, fill_value=0).round(3)
        return df

    df_all = make_df(counts_all, freqs_all)
    df_func = make_df(counts_func, freqs_func)
    df_hypo = make_df(counts_hypo, freqs_hypo)
    df_norm = make_df(counts_norm, freqs_norm)

    def add_er(df: pd.DataFrame) -> pd.DataFrame:
        er = []
        for t, f_val in zip(df["TYPE"], df["f"]):
            base = ref_freq.get(t, 0)
            er.append(0 if base == 0 else round(f_val / base, 3))
        df["ER"] = er
        return df

    df_func = add_er(df_func).sort_values("ER", ascending=False).reset_index(drop=True)
    df_hypo = add_er(df_hypo).sort_values("ER", ascending=False).reset_index(drop=True)
    df_norm = add_er(df_norm).sort_values("ER", ascending=False).reset_index(drop=True)

    def highlight_sets(df: pd.DataFrame) -> dict:
        nonzero = df[df["ER"] > 0]
        top_red = set(nonzero.head(10)["TYPE"].tolist())
        bottom_blue = set(nonzero.sort_values("ER", ascending=True).head(10)["TYPE"].tolist())
        return {"red": top_red, "blue": bottom_blue}

    return {
        "all": df_all,
        "path": df_func,
        "hypo": df_hypo,
        "benign": df_norm,
        "highlights": {
            "path": highlight_sets(df_func),
            "hypo": highlight_sets(df_hypo),
            "benign": highlight_sets(df_norm),
        },
        "counts": {
            "all": n_all,
            "path": n_func,
            "hypo": n_hypo,
            "benign": n_norm,
        },
    }


def _apply_panel_borders(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    thin = Side(style="thin", color="000000")
    thick = Side(style="thick", color="000000")
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            left = thin
            right = thin
            top = thin
            bottom = thin
            if row == min_row:
                top = thick
            if row == max_row:
                bottom = thick
            if col == min_col:
                left = thick
            if col == max_col:
                right = thick
            ws.cell(row=row, column=col).border = Border(
                left=left, right=right, top=top, bottom=bottom
            )


def _fill_panel(ws, min_row: int, max_row: int, min_col: int, max_col: int, color: str | None) -> None:
    if not color:
        return
    fill = PatternFill(fill_type="solid", fgColor=color)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row=row, column=col).fill = fill


def _write_panel_sheet(
    output_path: str,
    sheet_name: str,
    title: str,
    all_label: str,
    subgroup_label: str,
    panel_data: Dict[str, pd.DataFrame],
) -> None:
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    if output_file.exists():
        wb = load_workbook(output_file)
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        ws = wb.create_sheet(sheet_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:S1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, name="Calibri", size=10)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    counts = panel_data["counts"]

    ws.merge_cells("A2:C2")
    ws["A2"] = f"{all_label} All VUS (n = {counts['all']})"
    ws.merge_cells("E2:H2")
    ws["E2"] = f"{subgroup_label} VUS -> Functional Impact (n = {counts['path']})"
    ws.merge_cells("J2:M2")
    ws["J2"] = f"{subgroup_label} VUS -> Hypomorph (n = {counts['hypo']})"
    ws.merge_cells("O2:R2")
    ws["O2"] = f"{subgroup_label} VUS -> Normal function (n = {counts['benign']})"

    header_font = Font(bold=True, name="Calibri", size=10)
    for cell in ("A2", "E2", "J2", "O2"):
        ws[cell].font = header_font
        ws[cell].alignment = Alignment(horizontal="center", vertical="center")

    ws["A3"] = "TYPE"
    ws["B3"] = "counts"
    ws["C3"] = "f"
    ws["E3"] = "TYPE"
    ws["F3"] = "counts"
    ws["G3"] = "f"
    ws["H3"] = "ER (Funct Impact/all VUS)"
    ws["J3"] = "TYPE"
    ws["K3"] = "counts"
    ws["L3"] = "f"
    ws["M3"] = "ER (Hyp/all VUS)"
    ws["O3"] = "TYPE"
    ws["P3"] = "counts"
    ws["Q3"] = "f"
    ws["R3"] = "ER (Normal Funct/all VUS)"

    for col in ("A", "B", "C", "E", "F", "G", "J", "K", "L", "O", "P", "Q"):
        ws[f"{col}3"].font = header_font
        ws[f"{col}3"].alignment = Alignment(horizontal="center", vertical="center")
    for col in ("H", "M", "R"):
        ws[f"{col}3"].font = header_font
        ws[f"{col}3"].alignment = Alignment(horizontal="center", vertical="center", textRotation=90, wrap_text=True)

    ws.row_dimensions[3].height = 120

    def write_panel(df: pd.DataFrame, start_col: int, include_er: bool, red_set: set[str] | None, blue_set: set[str] | None):
        for i, (_, row) in enumerate(df.iterrows()):
            r = DATA_START_ROW + i
            ws.cell(row=r, column=start_col, value=row["TYPE"])
            ws.cell(row=r, column=start_col + 1, value=row["c"])
            ws.cell(row=r, column=start_col + 2, value=row["f"])
            if include_er:
                ws.cell(row=r, column=start_col + 3, value=row["ER"])

            type_cell = ws.cell(row=r, column=start_col)
            type_cell.alignment = Alignment(horizontal="left", vertical="center")
            if red_set and row["TYPE"] in red_set:
                type_cell.font = Font(name="Calibri", size=10, color="C00000")
            elif blue_set and row["TYPE"] in blue_set:
                type_cell.font = Font(name="Calibri", size=10, color="0070C0")
            else:
                type_cell.font = Font(name="Calibri", size=10, color="000000")

            for offset in range(1, 4 if include_er else 3):
                cell = ws.cell(row=r, column=start_col + offset)
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if offset == 1:
                    cell.number_format = "0"
                else:
                    cell.number_format = "0.000"
                cell.font = Font(name="Calibri", size=10, color="000000")

    highlights = panel_data.get("highlights", {})
    write_panel(panel_data["all"], 1, include_er=False, red_set=None, blue_set=None)
    write_panel(
        panel_data["path"],
        5,
        include_er=True,
        red_set=highlights.get("path", {}).get("red"),
        blue_set=highlights.get("path", {}).get("blue"),
    )
    write_panel(
        panel_data["hypo"],
        10,
        include_er=True,
        red_set=highlights.get("hypo", {}).get("red"),
        blue_set=highlights.get("hypo", {}).get("blue"),
    )
    write_panel(
        panel_data["benign"],
        15,
        include_er=True,
        red_set=highlights.get("benign", {}).get("red"),
        blue_set=highlights.get("benign", {}).get("blue"),
    )

    last_row = DATA_START_ROW + len(panel_data["all"]) - 1

    _fill_panel(ws, 2, last_row, 5, 8, PANEL2_FILL)
    _fill_panel(ws, 2, last_row, 10, 13, PANEL3_FILL)
    _fill_panel(ws, 2, last_row, 15, 18, PANEL4_FILL)

    _apply_panel_borders(ws, 2, last_row, 1, 3)
    _apply_panel_borders(ws, 2, last_row, 5, 8)
    _apply_panel_borders(ws, 2, last_row, 10, 13)
    _apply_panel_borders(ws, 2, last_row, 15, 18)

    for col_letter, width in {
        "A": 5.5,
        "B": 4.0,
        "C": 6.0,
        "D": 1.0,
        "E": 5.5,
        "F": 4.0,
        "G": 6.0,
        "H": 6.0,
        "I": 1.0,
        "J": 5.5,
        "K": 4.0,
        "L": 6.0,
        "M": 6.0,
        "N": 1.0,
        "O": 5.5,
        "P": 4.0,
        "Q": 6.0,
        "R": 6.0,
    }.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A4"
    wb.save(output_file)


def write_sup_table_14_15(
    brca1_table: pd.DataFrame,
    brca2_table: pd.DataFrame,
    brca1_ref_panel: pd.DataFrame,
    brca2_ref_panel: pd.DataFrame,
    sup12_df: pd.DataFrame,
    sup13_df: pd.DataFrame,
    output_path: str,
) -> None:
    def _ref_col(df: pd.DataFrame) -> str:
        try:
            return _find_col(df, ["T5", "Variant", "Variant ID", "variant", "variant id", "T5 (Variant)"])
        except KeyError:
            return str(df.columns[0])

    ref_col_b1 = _ref_col(brca1_ref_panel)
    ref_col_b2 = _ref_col(brca2_ref_panel)
    ref_brca1 = set(brca1_ref_panel[ref_col_b1].dropna().astype(str).tolist())
    ref_brca2 = set(brca2_ref_panel[ref_col_b2].dropna().astype(str).tolist())

    panel_brca1 = _panel_data_from_sup12(brca1_table, sup12_df, ref_brca1)
    panel_brca2 = _panel_data_from_sup12(brca2_table, sup13_df, ref_brca2)

    _write_panel_sheet(
        output_path,
        "Sup Table 14",
        "Supplementary Table 14: Frequency and enrichment ratios of BRCA1 amino acid substitution stratified by functional category",
        "BRCA1",
        "B1",
        panel_brca1,
    )
    _write_panel_sheet(
        output_path,
        "Sup Table 15",
        "Supplementary Table 15: Frequency and enrichment ratios of BRCA2 amino acid substitution stratified by functional category",
        "BRCA2",
        "B2",
        panel_brca2,
    )


AA3_TO_AA1 = {
    "Ala": "A", "Arg": "R", "Asn": "N", "Asp": "D", "Cys": "C",
    "Gln": "Q", "Glu": "E", "Gly": "G", "His": "H", "Ile": "I",
    "Leu": "L", "Lys": "K", "Met": "M", "Phe": "F", "Pro": "P",
    "Ser": "S", "Thr": "T", "Trp": "W", "Tyr": "Y", "Val": "V",
}


def norm_col(c: str) -> str:
    return re.sub(r"\s+", " ", str(c).strip().lower())


def find_col(df: pd.DataFrame, candidates: Iterable[str]) -> str:
    """
    Find a column in df matching any candidate name (case/space-insensitive).
    Raises KeyError if none found.
    """
    mapping = {norm_col(c): c for c in df.columns}
    for cand in candidates:
        key = norm_col(cand)
        if key in mapping:
            return mapping[key]
    # fuzzy contains
    for cand in candidates:
        key = norm_col(cand)
        for k_norm, k_orig in mapping.items():
            if key in k_norm:
                return k_orig
    raise KeyError(f"Could not find any of columns: {list(candidates)}")


def resolve_sheet_name(xlsx_path: str, preferred: str, fallbacks: List[str]) -> str:
    """
    Return a sheet name that exists in the workbook.
    Tries preferred; else tries case-insensitive exact match; else tries regex-like fallbacks.
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    names = wb.sheetnames
    if preferred in names:
        return preferred

    lower_map = {n.lower(): n for n in names}
    if preferred.lower() in lower_map:
        return lower_map[preferred.lower()]

    # try fallbacks as "contains" patterns (case-insensitive)
    for pat in fallbacks:
        pat_l = pat.lower()
        for n in names:
            if pat_l in n.lower():
                return n

    raise ValueError(
        f"Sheet not found. preferred={preferred!r}. "
        f"Available sheets: {names}"
    )


def extract_type(protein_change: str) -> Optional[str]:
    """
    From "p.A123T" or "A123T" -> "AT"
    Also supports three-letter like "p.Ala123Thr" -> "AT"
    Returns None if not parseable.
    """
    if protein_change is None or (isinstance(protein_change, float) and pd.isna(protein_change)):
        return None
    s = str(protein_change).strip()

    # one-letter forms
    m = re.match(r"^(?:p\.)?([A-Z])\d+([A-Z])$", s)
    if m:
        return m.group(1) + m.group(2)

    # three-letter forms (HGVS-like)
    m = re.match(r"^(?:p\.)?([A-Z][a-z]{2})\d+([A-Z][a-z]{2})$", s)
    if m:
        a1 = AA3_TO_AA1.get(m.group(1))
        a2 = AA3_TO_AA1.get(m.group(2))
        if a1 and a2:
            return a1 + a2

    return None


def read_sup_table(
    xlsx_path: str,
    sheet_name: str,
    header_row_index: int = 1,  # row 2 in Excel => header=1 in pandas
) -> pd.DataFrame:
    df = pd.read_excel(xlsx_path, sheet_name=sheet_name, header=header_row_index, engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]
    return df


@dataclass
class Block:
    title: str
    side_label: str
    fill_hex: str
    data: pd.DataFrame              # columns: TYPE, counts, f, [ER]
    highlight_types: Set[str]       # TYPE values to render in red


def make_counts_freq(
    variants: Iterable[str],
    n_total: int,
) -> pd.DataFrame:
    s = pd.Series(list(variants), name="T5").dropna()
    types = s.map(extract_type).dropna()
    counts = types.value_counts()
    df = counts.rename("counts").to_frame()
    df["f"] = (df["counts"] / float(n_total)).round(3)
    df.index.name = "TYPE"
    df = df.reset_index()
    return df


def add_enrichment(df_sub: pd.DataFrame, f_ref: Dict[str, float], er_colname: str) -> pd.DataFrame:
    df = df_sub.copy()
    df[er_colname] = df["TYPE"].map(lambda t: (df.loc[df["TYPE"] == t, "f"].iloc[0] / f_ref.get(t, float("nan"))))
    df[er_colname] = df[er_colname].replace([float("inf"), -float("inf")], pd.NA).round(3)
    return df


def top_n_types_by_er(df: pd.DataFrame, er_col: str, n: int = 10) -> Set[str]:
    tmp = df.dropna(subset=[er_col]).copy()
    tmp = tmp[tmp["counts"] > 0]
    tmp = tmp.sort_values(er_col, ascending=False).head(n)
    return set(tmp["TYPE"].tolist())


def thin_border() -> Border:
    side = Side(style="thin", color="999999")
    return Border(left=side, right=side, top=side, bottom=side)


def write_block(
    ws,
    start_col: int,
    start_row: int,
    block: Block,
    has_er: bool,
    max_data_rows: int,
):
    """
    Layout:
      Row start_row: merged block title across data columns
      Row start_row+1: column headers
      Row start_row+2.. : data rows
      plus one narrow "side label" column to the right, merged vertically across the block
    """
    title_row = start_row
    header_row = start_row + 1
    data_row0 = start_row + 2

    # columns
    # data columns
    cols = ["TYPE", "counts", "f"] + (["ER"] if has_er else [])
    n_data_cols = len(cols)
    side_col = start_col + n_data_cols  # one extra col

    fill = PatternFill("solid", fgColor=block.fill_hex)
    bdr = thin_border()

    # title merged cell
    ws.merge_cells(
        start_row=title_row, start_column=start_col,
        end_row=title_row, end_column=start_col + n_data_cols - 1
    )
    c = ws.cell(row=title_row, column=start_col, value=block.title)
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center")
    # fill title row across data columns
    for j in range(start_col, start_col + n_data_cols):
        cell = ws.cell(row=title_row, column=j)
        cell.fill = fill
        cell.border = bdr

    # headers
    for j, name in enumerate(cols, start=start_col):
        cell = ws.cell(row=header_row, column=j, value=name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = fill
        cell.border = bdr

    # side label (rotated)
    ws.merge_cells(
        start_row=title_row, start_column=side_col,
        end_row=data_row0 + max_data_rows - 1, end_column=side_col
    )
    side_cell = ws.cell(row=title_row, column=side_col, value=block.side_label)
    side_cell.font = Font(bold=True)
    side_cell.alignment = Alignment(horizontal="center", vertical="center", text_rotation=90, wrap_text=True)
    side_cell.fill = fill
    side_cell.border = bdr

    # paint background for side label merged region
    for r in range(title_row, data_row0 + max_data_rows):
        cell = ws.cell(row=r, column=side_col)
        cell.fill = fill
        cell.border = bdr

    # write data (pad to max_data_rows for consistent block height)
    df = block.data.copy()

    # normalize ER column name to "ER" for writing if present
    if has_er:
        # the ER column is the 4th column in df
        er_actual = [c for c in df.columns if c not in ("TYPE", "counts", "f")][0]
        df = df.rename(columns={er_actual: "ER"})

    # ensure column order
    df = df[cols]

    for i in range(max_data_rows):
        r = data_row0 + i
        if i < len(df):
            row = df.iloc[i]
            vals = [row.get("TYPE"), row.get("counts"), row.get("f")] + ([row.get("ER")] if has_er else [])
        else:
            vals = [None] * n_data_cols

        for j, v in enumerate(vals, start=start_col):
            cell = ws.cell(row=r, column=j, value=v)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = fill
            cell.border = bdr
            if j == start_col and v is not None:  # TYPE col
                if str(v) in block.highlight_types:
                    cell.font = Font(bold=True, color="FF0000")
                else:
                    cell.font = Font(bold=True)

    # set column widths
    widths = {
        "TYPE": 6,
        "counts": 8,
        "f": 8,
        "ER": 14,
    }
    for j, name in enumerate(cols, start=start_col):
        ws.column_dimensions[get_column_letter(j)].width = widths.get(name, 10)
    ws.column_dimensions[get_column_letter(side_col)].width = 4


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook", help="Input .xlsx containing Sup Tables 1/3/5/12")
    ap.add_argument("-o", "--out", default="SupTable14.xlsx", help="Output .xlsx")
    ap.add_argument("--sup1", default="Sup Table 1")
    ap.add_argument("--sup3", default="Sup Table 3")
    ap.add_argument("--sup5", default="Sup Table 5")
    ap.add_argument("--sup12", default="Sup Table 12")
    ap.add_argument("--top-n-red", type=int, default=10, help="Highlight top-N TYPEs by ER in red per subset block")
    args = ap.parse_args()

    # Resolve sheet names (robust to minor naming differences)
    sup1_sheet = resolve_sheet_name(args.workbook, args.sup1, ["sup table 1", "table 1", "brca1_table"])
    sup3_sheet = resolve_sheet_name(args.workbook, args.sup3, ["sup table 3", "table 3", "metadata"])
    sup5_sheet = resolve_sheet_name(args.workbook, args.sup5, ["sup table 5", "table 5", "reference"])
    sup12_sheet = resolve_sheet_name(args.workbook, args.sup12, ["sup table 12", "table 12", "functional evidence"])

    sup1 = read_sup_table(args.workbook, sup1_sheet, header_row_index=1)
    sup3 = read_sup_table(args.workbook, sup3_sheet, header_row_index=1)
    sup5 = read_sup_table(args.workbook, sup5_sheet, header_row_index=1)
    sup12 = read_sup_table(args.workbook, sup12_sheet, header_row_index=1)

    # Columns (Sup1)
    col_t5_1 = find_col(sup1, ["T5"])
    col_t6_1 = find_col(sup1, ["T6"])
    # Columns (Sup5)
    col_t5_5 = find_col(sup5, ["T5"])
    col_t6_5 = find_col(sup5, ["T6"])
    # Columns (Sup12)
    col_t5_12 = find_col(sup12, ["T5"])
    col_t6_12 = find_col(sup12, ["T6"])
    col_fe = find_col(sup12, ["Functional evidence in favor of"])
    col_hypo = find_col(sup12, ["Hypomorph observation"])
    # Optional, if present: integrated criteria (useful if your file uses PS3/BS3 strings)
    col_criteria = None
    try:
        col_criteria = find_col(sup12, ["Integrated ACMG evidence criteria"])
    except KeyError:
        pass

    # -----------------------------
    # Consistency checks using Sup3
    # -----------------------------
    # Compare track ids in Sup1 (T8+) vs Sup3 track list (if we can infer it).
    track_cols_sup1 = [c for c in sup1.columns if re.fullmatch(r"T\d+", str(c)) and int(str(c)[1:]) >= 8]
    track_ids_sup1 = set(track_cols_sup1)

    track_ids_sup3 = set()
    # Try to detect a track-id column in sup3 by scanning for values like "T8"
    for c in sup3.columns:
        s = sup3[c].astype(str).str.strip()
        if (s.str.match(r"^T\d+$").sum() >= 5):  # heuristic
            track_ids_sup3 = set(s[s.str.match(r"^T\d+$")].tolist())
            break

    if track_ids_sup3:
        missing_in_meta = sorted(track_ids_sup1 - track_ids_sup3)
        if missing_in_meta:
            print(f"[WARN] Tracks present in Sup1 but not found in Sup3 metadata: {missing_in_meta[:10]}{'...' if len(missing_in_meta)>10 else ''}")

    # -----------------------------
    # Define VUS universe (Sup1) and exclude reference panel (Sup5)
    # -----------------------------
    ref_panel = sup5[[col_t5_5, col_t6_5]].dropna(subset=[col_t5_5]).copy()
    ref_variants: Set[str] = set(ref_panel[col_t5_5].astype(str).tolist())

    sup1_min = sup1[[col_t5_1, col_t6_1]].dropna(subset=[col_t5_1]).copy()
    sup1_min[col_t5_1] = sup1_min[col_t5_1].astype(str)

    # VUS: T6 empty/NaN, AND not in reference panel
    vus_sup1 = sup1_min[sup1_min[col_t6_1].isna()].copy()
    vus_sup1 = vus_sup1[~vus_sup1[col_t5_1].isin(ref_variants)]
    vus_variants: Set[str] = set(vus_sup1[col_t5_1].tolist())

    # -----------------------------
    # Join to Sup12 for functional stratification
    # -----------------------------
    sup12_min = sup12[[col_t5_12, col_t6_12, col_fe, col_hypo] + ([col_criteria] if col_criteria else [])].copy()
    sup12_min = sup12_min.dropna(subset=[col_t5_12])
    sup12_min[col_t5_12] = sup12_min[col_t5_12].astype(str)

    # Filter to VUS universe from Sup1 (ensures table 14 is based on Sup1-defined VUS)
    sup12_vus = sup12_min[sup12_min[col_t5_12].isin(vus_variants)].copy()

    fe = sup12_vus[col_fe].astype(str)
    hypo_obs = sup12_vus[col_hypo].astype(str)

    is_hypo = hypo_obs.str.upper().eq("Y") | fe.str.contains("hypomorph", case=False, na=False)

    # These rules match the typical Table 12 semantics:
    #   - Functional Impact: functional evidence favors Pathogenic (PS3-like), excluding hypomorph
    #   - Normal function: functional evidence favors Benign (BS3-like), excluding hypomorph
    is_path = fe.str.contains("pathogenic", case=False, na=False)
    is_ben = fe.str.contains("benign", case=False, na=False)

    if col_criteria:
        crit = sup12_vus[col_criteria].astype(str)
        is_path = is_path | crit.str.contains("ps3", case=False, na=False)
        is_ben = is_ben | crit.str.contains("bs3", case=False, na=False)

    func_impact_vars = set(sup12_vus.loc[is_path & ~is_hypo, col_t5_12].tolist())
    hypomorph_vars = set(sup12_vus.loc[is_hypo, col_t5_12].tolist())
    normal_vars = set(sup12_vus.loc[is_ben & ~is_hypo, col_t5_12].tolist())

    # -----------------------------
    # Compute frequency and enrichment
    # -----------------------------
    n_all = len(sup12_vus)
    n_func = len(func_impact_vars)
    n_hypo = len(hypomorph_vars)
    n_norm = len(normal_vars)

    if n_all == 0:
        raise RuntimeError("No VUS variants found after filtering. Check sheet names and T5/T6 columns.")

    df_all = make_counts_freq(sup12_vus[col_t5_12].tolist(), n_all)
    df_all = df_all.sort_values("TYPE").reset_index(drop=True)
    f_ref = dict(zip(df_all["TYPE"], df_all["f"]))

    df_func = make_counts_freq(func_impact_vars, n_func) if n_func else pd.DataFrame(columns=["TYPE","counts","f"])
    df_hypo = make_counts_freq(hypomorph_vars, n_hypo) if n_hypo else pd.DataFrame(columns=["TYPE","counts","f"])
    df_norm = make_counts_freq(normal_vars, n_norm) if n_norm else pd.DataFrame(columns=["TYPE","counts","f"])

    df_func = add_enrichment(df_func, f_ref, "ER (Func Impact/All VUS)") if len(df_func) else df_func
    df_hypo = add_enrichment(df_hypo, f_ref, "ER (Hyp/All VUS)") if len(df_hypo) else df_hypo
    df_norm = add_enrichment(df_norm, f_ref, "ER (Normal/All VUS)") if len(df_norm) else df_norm

    # Sort subset blocks by ER desc to mimic screenshot-style emphasis
    if len(df_func):
        df_func = df_func.sort_values("ER (Func Impact/All VUS)", ascending=False).reset_index(drop=True)
    if len(df_hypo):
        df_hypo = df_hypo.sort_values("ER (Hyp/All VUS)", ascending=False).reset_index(drop=True)
    if len(df_norm):
        df_norm = df_norm.sort_values("ER (Normal/All VUS)", ascending=False).reset_index(drop=True)

    hi_func = top_n_types_by_er(df_func, "ER (Func Impact/All VUS)", n=args.top_n_red) if len(df_func) else set()
    hi_hypo = top_n_types_by_er(df_hypo, "ER (Hyp/All VUS)", n=args.top_n_red) if len(df_hypo) else set()
    hi_norm = top_n_types_by_er(df_norm, "ER (Normal/All VUS)", n=args.top_n_red) if len(df_norm) else set()

    # -----------------------------
    # Build formatted Excel output
    # -----------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Sup Table 14"

    # Title row
    total_cols = 4 + 5 + 5 + 5  # block1:3+side, others:4+side
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title = "Supplementary Table 14: Frequency and enrichment ratios of BRCA1 amino acid substitution stratified by functional category"
    tcell = ws.cell(row=1, column=1, value=title)
    tcell.font = Font(bold=True)
    tcell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    blocks: List[Tuple[Block, bool]] = [
        (
            Block(
                title=f"BRCA1 All VUS (n = {n_all})",
                side_label="All VUS",
                fill_hex="E6F3E1",  # light green
                data=df_all[["TYPE", "counts", "f"]],
                highlight_types=set(),
            ),
            False,
        ),
        (
            Block(
                title=f"B1 VUS -> Functional Impact (n = {n_func})",
                side_label="Func\nImpact\nVUS",
                fill_hex="F7DED6",  # light pink
                data=df_func,
                highlight_types=hi_func,
            ),
            True,
        ),
        (
            Block(
                title=f"B1 VUS -> Hypomorph (n = {n_hypo})",
                side_label="Hyp/all\nVUS",
                fill_hex="FBE9C8",  # light yellow
                data=df_hypo,
                highlight_types=hi_hypo,
            ),
            True,
        ),
        (
            Block(
                title=f"B1 VUS -> Normal function (n = {n_norm})",
                side_label="ER\n(Normal\nFunc/\nAll VUS)",
                fill_hex="DDEBFA",  # light blue
                data=df_norm,
                highlight_types=hi_norm,
            ),
            True,
        ),
    ]

    max_rows = max(len(b.data) for b, _ in blocks)
    start_row = 2

    # Block start columns:
    #  - Block1 uses 3 data cols + 1 side = 4 cols
    #  - Blocks2-4 use 4 data cols + 1 side = 5 cols each
    col = 1
    for blk, has_er in blocks:
        write_block(
            ws=ws,
            start_col=col,
            start_row=start_row,
            block=blk,
            has_er=has_er,
            max_data_rows=max_rows,
        )
        col += (4 if not has_er else 5)

    # General sheet tweaks
    ws.freeze_panes = ws["A4"]
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 18

    print(f"[INFO] n_all={n_all}, n_func={n_func}, n_hypo={n_hypo}, n_norm={n_norm}")
    wb.save(args.out)
    print(f"[OK] Wrote: {args.out}")


if __name__ == "__main__":
    main()
