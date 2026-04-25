from pathlib import Path
import tempfile
from typing import Dict, List, Tuple

import math
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font
from sup_table_12_13 import write_sup_table_12_13
from sup_table_18_19 import write_sup_tables_18_19

DATA_START_COL = "T8"
REFERENCE_COL = "T6"
DOCUMENTED_COL = "T7"

BRCA1_TOTAL_MISSENSE = 11009
BRCA2_TOTAL_MISSENSE = 20169
EXCLUDED_T6_VARIANTS = {"p.M1I", "p.M1K", "p.M1R", "p.M1T", "p.M1V"}
Z_SCORE = 1.95996

INPUT_PATH = "dataset/SUPP_TABLES_BRCA12_APR_2026.xlsx"
OUTPUT_PATH = "sup7_sup8_new_cal.xlsx"
ALPHA_PATH = "dataset/AlphaMissense_Calculations_all.xlsx"
OTHER_POINTS_PATH = "dataset/ACMG_other_points.xlsx"

SHEETS = {
    "Sup Table 1": "BRCA1_table",
    "Sup Table 2": "BRCA2_table",
    "Sup Table 3": "BRCA1_metadata",
    "Sup Table 4": "BRCA2_metadata",
}


def load_tables(path: str) -> dict:
    tables = {}
    for sheet_name, var_name in SHEETS.items():
        tables[var_name] = pd.read_excel(path, sheet_name=sheet_name, header=1)
    return tables


def _resolve_existing_path(*candidates: str) -> str | None:
    for candidate in candidates:
        if not candidate:
            continue
        if Path(candidate).exists():
            return candidate
    return None


def _prepare_other_points_workbook(path: str) -> str:
    wb = load_workbook(path, read_only=True, data_only=True)
    sheetnames = set(wb.sheetnames)
    if {"BRCA1", "BRCA2"}.issubset(sheetnames):
        return path
    if not {"BRCA1_evidence_criteria", "BRCA2_evidence_criteria"}.issubset(sheetnames):
        return path

    tmp = tempfile.NamedTemporaryFile(
        prefix="normalized_other_points_",
        suffix=".xlsx",
        delete=False,
    )
    tmp_path = tmp.name
    tmp.close()

    brca1_df = pd.read_excel(path, sheet_name="BRCA1_evidence_criteria")
    brca2_df = pd.read_excel(path, sheet_name="BRCA2_evidence_criteria")
    with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
        brca1_df.to_excel(writer, sheet_name="BRCA1", index=False)
        brca2_df.to_excel(writer, sheet_name="BRCA2", index=False)
    return tmp_path


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = out.columns.astype(str).str.strip()
    return out


def _resolve_track_col(meta_df: pd.DataFrame) -> str:
    meta_df = _normalize_columns(meta_df)
    norm_map = {str(c).strip().lower(): c for c in meta_df.columns}
    candidates = [
        "track #",
        "track#",
        "track number",
        "track no",
        "track id",
        "track",
    ]
    for candidate in candidates:
        if candidate in norm_map:
            return norm_map[candidate]
    return meta_df.columns[0]


def _track_whitelist(meta_df: pd.DataFrame) -> set[str]:
    if meta_df is None or meta_df.empty:
        return set()
    track_col = _resolve_track_col(meta_df)
    whitelist: set[str] = set()
    for val in meta_df[track_col].dropna():
        track_raw = str(val).strip()
        if not track_raw or track_raw.lower() == "nan":
            continue
        if track_raw.startswith("T"):
            whitelist.add(track_raw)
        else:
            whitelist.add(f"T{track_raw}")
    return whitelist


def _filter_data_cols(data_df: pd.DataFrame, meta_df: pd.DataFrame | None) -> pd.DataFrame:
    whitelist = _track_whitelist(meta_df)
    if not whitelist:
        return data_df
    keep = [c for c in data_df.columns if str(c).strip() in whitelist]
    return data_df[keep]


def _norm_t6(series: pd.Series) -> pd.Series:
    def _norm(val: object) -> str:
        if pd.isna(val):
            return ""
        text = str(val).strip().replace(" ", "")
        if text.endswith(".0"):
            text = text[:-2]
        return text

    return series.map(_norm)


def _get_data_block(
    df: pd.DataFrame,
    meta_df: pd.DataFrame | None = None,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    df = _normalize_columns(df)
    if DATA_START_COL not in df.columns:
        raise KeyError(f"Missing data start column: {DATA_START_COL}")
    if REFERENCE_COL not in df.columns:
        raise KeyError(f"Missing reference column: {REFERENCE_COL}")
    if DOCUMENTED_COL not in df.columns:
        raise KeyError(f"Missing documented column: {DOCUMENTED_COL}")
    start_idx = df.columns.get_loc(DATA_START_COL)
    data_df = df.iloc[:, start_idx:]
    data_df = _filter_data_cols(data_df, meta_df)
    return df, data_df


def _resolve_metadata_columns(meta_df: pd.DataFrame) -> Tuple[str, str]:
    meta_df = _normalize_columns(meta_df)
    norm_map = {str(c).strip().lower(): c for c in meta_df.columns}

    track_num_candidates = [
        "track #",
        "track#",
        "track number",
        "track no",
        "track id",
    ]
    track_name_candidates = ["track", "assay", "track name", "assay name"]

    track_num_col = next((norm_map[c] for c in track_num_candidates if c in norm_map), None)
    if track_num_col is None:
        track_num_col = meta_df.columns[0]

    track_name_col = next(
        (
            norm_map[c]
            for c in track_name_candidates
            if c in norm_map and norm_map[c] != track_num_col
        ),
        None,
    )
    if track_name_col is None:
        track_name_col = meta_df.columns[1] if len(meta_df.columns) > 1 else track_num_col

    return track_num_col, track_name_col


def _build_track_map(meta_df: pd.DataFrame) -> Dict[str, str]:
    track_num_col, track_name_col = _resolve_metadata_columns(meta_df)
    track_map: Dict[str, str] = {}

    for _, row in meta_df.iterrows():
        track_raw = str(row[track_num_col]).strip()
        if not track_raw or track_raw.lower() == "nan":
            continue
        track_name = str(row[track_name_col]).strip() if track_name_col in row else ""
        track_map[track_raw] = track_name
        if track_raw.startswith("T"):
            track_map[track_raw[1:]] = track_name
        else:
            track_map[f"T{track_raw}"] = track_name

    return track_map


def _custom_ci(p: float, n: int, z: float) -> Tuple[float, float]:
    if n == 0:
        return 0.0, 0.0
    q = 1 - p
    z2 = z * z
    denom = 2 * (n + z2)

    inner_lower = z2 - 2 - (1 / n) + 4 * p * ((n * q) + 1)
    inner_upper = z2 + 2 - (1 / n) + 4 * p * ((n * q) - 1)
    inner_lower = max(inner_lower, 0.0)
    inner_upper = max(inner_upper, 0.0)

    numerador_lower = (2 * n * p) + z2 - 1 - (z * math.sqrt(inner_lower))
    numerador_upper = (2 * n * p) + z2 + 1 + (z * math.sqrt(inner_upper))

    lower = numerador_lower / denom
    upper = numerador_upper / denom
    return max(0.0, lower), min(1.0, upper)


def _classify_odds(odds: float) -> str:
    if 0.0001 < odds < 0.0029:
        return "BS3"
    if odds < 0.053:
        return "BS3"
    if odds < 0.23:
        return "BS3_moderate"
    if odds < 0.48:
        return "BS3_supporting"
    if odds > 350:
        return "PS3_very_strong"
    if odds > 18.7:
        return "PS3"
    if odds > 4.3:
        return "PS3_moderate"
    if odds > 2.1:
        return "PS3_supporting"
    return "Indeterminate"


def _safe_odds(p2: float, p1: float) -> float:
    denom = (1 - p2) * p1
    if denom == 0:
        return 0.0
    return (p2 * (1 - p1)) / denom


def build_track_rows(
    brca_table: pd.DataFrame,
    metadata: pd.DataFrame,
    total_missense: int,
) -> List[List]:
    df, data_df = _get_data_block(brca_table, metadata)
    track_map = _build_track_map(metadata)
    track_cols = list(data_df.columns)

    t6_clean = _norm_t6(df[REFERENCE_COL])
    if "T5" in df.columns:
        mask = df["T5"].astype(str).str.strip().isin(EXCLUDED_T6_VARIANTS)
        if mask.any():
            t6_clean = t6_clean.mask(mask, "")
    t7 = pd.to_numeric(df[DOCUMENTED_COL], errors="coerce")
    ref_mask = t6_clean.ne("")
    path_mask = t6_clean.isin({"4", "5", "4;5"})
    ben_mask = t6_clean.isin({"1", "2", "1;2"})

    rows = []
    for track_col in track_cols:
        track_id = str(track_col).strip()
        track_name = track_map.get(track_id, "")

        col_vals = df[track_col]
        present = col_vals.notna()
        tested_count = int(present.sum())
        tested_pct = 0.0 if total_missense == 0 else round((tested_count / total_missense) * 100, 2)

        documented_classified = int((present & t7.eq(1)).sum())
        doc_pct = 0.0 if tested_count == 0 else round((documented_classified / tested_count) * 100, 2)

        total_ref = int((present & ref_mask).sum())
        path_ref = int((present & path_mask).sum())
        ben_ref = int((present & ben_mask).sum())

        vals_num = pd.to_numeric(col_vals, errors="coerce")
        tested_mask = present & vals_num.notna()
        tp = int((tested_mask & vals_num.eq(2) & path_mask).sum())
        fn = int((tested_mask & vals_num.isin([0, 1]) & path_mask).sum())
        tn = int((tested_mask & vals_num.eq(0) & ben_mask).sum())
        fp = int((tested_mask & vals_num.isin([1, 2]) & ben_mask).sum())

        sens_denom = tp + fn
        sensitivity = 0.0 if sens_denom == 0 else tp / sens_denom
        sens_low, sens_high = _custom_ci(sensitivity, sens_denom, Z_SCORE)

        spec_denom = tn + fp
        specificity = 0.0 if spec_denom == 0 else tn / spec_denom
        spec_low, spec_high = _custom_ci(specificity, spec_denom, Z_SCORE)

        # The special recalculation excludes false positives / false negatives
        # from the assay-result block before the downstream P2 and odds steps.
        func_abnormal = tp
        ind_ref_path = int((tested_mask & vals_num.eq(1) & path_mask).sum())
        ind_ref_benign = int((tested_mask & vals_num.eq(1) & ben_mask).sum())
        func_normal = tn

        p1_denom = tp + tn + fp + fn
        p1_raw = 0.0 if p1_denom == 0 else (tp + fn) / p1_denom
        p2_path_raw = func_abnormal / (func_abnormal + 0.5)
        p2_benign_raw = 0.5 / (func_normal + 0.5)
        oddspath_path_raw = _safe_odds(p2_path_raw, p1_raw)
        oddspath_benign_raw = _safe_odds(p2_benign_raw, p1_raw)

        p1 = round(p1_raw, 2)
        p2_path = round(p2_path_raw, 2)
        p2_benign = round(p2_benign_raw, 2)
        oddspath_path = round(oddspath_path_raw, 7)
        oddspath_benign = round(oddspath_benign_raw, 7)
        acmg_path = "N/A" if oddspath_path == 0 else _classify_odds(oddspath_path)
        acmg_benign = "N/A" if oddspath_benign == 0 else _classify_odds(oddspath_benign)

        rows.append(
            [
                track_id,
                track_name,
                tested_count,
                tested_pct,
                documented_classified,
                doc_pct,
                total_ref,
                path_ref,
                ben_ref,
                round(sensitivity, 2),
                round(sens_low, 2),
                round(sens_high, 2),
                round(specificity, 2),
                round(spec_low, 2),
                round(spec_high, 2),
                p1,
                func_abnormal,
                ind_ref_path,
                ind_ref_benign,
                func_normal,
                p2_path,
                p2_benign,
                oddspath_path,
                oddspath_benign,
                acmg_path,
                acmg_benign,
            ]
        )

    return rows


def build_track_classification_map(
    brca_table: pd.DataFrame,
    metadata: pd.DataFrame,
    total_missense: int,
) -> Dict[str, Dict[str, str]]:
    rows = build_track_rows(brca_table, metadata, total_missense)
    class_map: Dict[str, Dict[str, str]] = {}
    for row in rows:
        track_id = str(row[0]).strip()
        class_map[track_id] = {"path": row[-2], "benign": row[-1]}
    return class_map


def write_sup_table(
    brca_table: pd.DataFrame,
    metadata: pd.DataFrame,
    output_path: str,
    sheet_name: str,
    title: str,
    total_missense: int,
) -> None:
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    if output_file.exists():
        wb = load_workbook(output_file)
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        ws = wb.create_sheet(sheet_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

    ws["A1"] = title
    ws["A1"].font = Font(bold=True)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    ws.merge_cells("G2:I2")
    ws["G2"] = "# Classified (Ref panel)"
    ws["G2"].font = Font(bold=True)

    ws.merge_cells("J2:O2")
    ws["J2"] = "Specificity and Sensitivity"
    ws["J2"].font = Font(bold=True)

    ws.merge_cells("P2:P3")
    ws["P2"] = "Prior probability (P1)"
    ws["P2"].font = Font(bold=True)
    ws["P2"].alignment = Alignment(textRotation=90, horizontal="center", vertical="center")

    ws.merge_cells("Q2:T2")
    ws["Q2"] = "Assay result"
    ws["Q2"].font = Font(bold=True)

    ws.merge_cells("U2:V2")
    ws["U2"] = "+0.5 proportions (Posterior, P2)"
    ws["U2"].font = Font(bold=True)

    ws.merge_cells("W2:X2")
    ws["W2"] = "Odds Path"
    ws["W2"].font = Font(bold=True)

    ws.merge_cells("Y2:Z2")
    ws["Y2"] = "ASSAY CREDENTIALS"
    ws["Y2"].font = Font(bold=True)

    headers = [
        "Track",
        "Assay",
        "# of missense variants tested",
        f"% of all possible missense (n = {total_missense}) tested",
        "# of all documented missense variants classified",
        "% of all documented missense variants classified",
        "Total",
        "Path",
        "Benign",
        "Sensitivity",
        "95% CI lower",
        "95% CI upper",
        "Specificity",
        "95% CI lower",
        "95% CI upper",
        "",
        "Functionally abnormal",
        "Indeterminate (ref path)",
        "Indeterminate (ref benign)",
        "Functionally normal",
        "Path",
        "Benign",
        "Path",
        "Benign",
        "Path",
        "Benign",
    ]

    rotate_headers = {"G", "H", "I", "J", "K", "L", "M", "N", "O"}
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        if isinstance(cell, MergedCell):
            continue
        cell.value = header
        cell.font = Font(bold=True)
        if cell.column_letter in rotate_headers:
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                textRotation=90,
                wrap_text=True,
            )

    rows = build_track_rows(brca_table, metadata, total_missense)
    start_row = 4
    for r_idx, row in enumerate(rows):
        ws.cell(row=start_row + r_idx, column=1, value=row[0])
        ws.cell(row=start_row + r_idx, column=2, value=row[1])
        for c_idx in range(2, len(row)):
            ws.cell(row=start_row + r_idx, column=c_idx + 1, value=row[c_idx])

    for r in range(start_row, start_row + len(rows)):
        ws[f"W{r}"].number_format = "0.0000000"
        ws[f"X{r}"].number_format = "0.0000000"

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            cell.font = Font(name="Arial", size=10, bold=bool(cell.font.bold))
            if cell.column_letter == "A":
                if cell.row == 1:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
                elif cell.row >= 3:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            elif cell.column_letter == "B" and cell.row >= 4:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            elif cell.column_letter == "P" and cell.row in (2, 3):
                cell.alignment = Alignment(horizontal="center", vertical="center", textRotation=90, wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    red_font = Font(name="Arial", size=10, bold=False, color="FF0000")
    for r in range(start_row, start_row + len(rows)):
        cell_y = ws[f"Y{r}"]
        val_y = "" if cell_y.value is None else str(cell_y.value)
        if "BS3" in val_y:
            cell_y.value = "Indeterminate*"
            cell_y.font = red_font

        cell_z = ws[f"Z{r}"]
        val_z = "" if cell_z.value is None else str(cell_z.value)
        if "PS3" in val_z:
            cell_z.value = "Indeterminate*"
            cell_z.font = red_font

    for col_letter in (
        "A",
        "C",
        "D",
        "E",
        "F",
        "G",
        "H",
        "I",
        "J",
        "K",
        "L",
        "M",
        "N",
        "O",
        "P",
        "Q",
        "R",
        "S",
        "T",
        "U",
        "V",
        "W",
        "X",
        "Y",
        "Z",
    ):
        ws.column_dimensions[col_letter].width = 14

    max_len_b = 0
    for cell in ws["B"]:
        if cell.value is None:
            continue
        max_len_b = max(max_len_b, len(str(cell.value)))
    if max_len_b > 0:
        ws.column_dimensions["B"].width = max(14, min(max_len_b + 2, 120))

    for row_idx in range(4, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 11.25

    wb.save(output_file)


def generate_workbook(
    input_path: str = INPUT_PATH,
    output_path: str = OUTPUT_PATH,
    alpha_path: str = ALPHA_PATH,
    other_points_path: str = OTHER_POINTS_PATH,
) -> None:
    tables = load_tables(input_path)
    write_sup_table(
        tables["BRCA1_table"],
        tables["BRCA1_metadata"],
        output_path,
        "Sup Table 7",
        "Supplementary Table 7: Summary statistics, specificity, sensitivity, odds of pathogenicity, and likelihood ratios for each BRCA1 track",
        BRCA1_TOTAL_MISSENSE,
    )
    write_sup_table(
        tables["BRCA2_table"],
        tables["BRCA2_metadata"],
        output_path,
        "Sup Table 8",
        "Supplementary Table 8: Summary statistics, specificity, sensitivity, odds of pathogenicity, and likelihood ratios for each BRCA2 track",
        BRCA2_TOTAL_MISSENSE,
    )
    brca1_class_map = build_track_classification_map(
        tables["BRCA1_table"],
        tables["BRCA1_metadata"],
        BRCA1_TOTAL_MISSENSE,
    )
    brca2_class_map = build_track_classification_map(
        tables["BRCA2_table"],
        tables["BRCA2_metadata"],
        BRCA2_TOTAL_MISSENSE,
    )
    sup12_df, sup13_df = write_sup_table_12_13(
        tables["BRCA1_table"],
        tables["BRCA2_table"],
        brca1_class_map,
        brca2_class_map,
        output_path,
        tables["BRCA1_metadata"],
        tables["BRCA2_metadata"],
    )
    resolved_alpha_path = _resolve_existing_path(
        alpha_path,
        "dataset/AlphaMissense_Calculations_all.xlsx",
        "AlphaMissense_Calculations_all.xlsx",
    )
    resolved_other_points_path = _resolve_existing_path(
        other_points_path,
        "dataset/ACMG_other_points.xlsx",
        "dataset/evidence_criteria_v6_BRCA12.xlsx",
        "ACMG_other_points.xlsx",
        "evidence_criteria_v6_BRCA12.xlsx",
    )
    if resolved_alpha_path and resolved_other_points_path:
        prepared_other_points_path = _prepare_other_points_workbook(resolved_other_points_path)
        write_sup_tables_18_19(
            sup12_df,
            sup13_df,
            resolved_alpha_path,
            prepared_other_points_path,
            output_path,
        )
    else:
        missing = []
        if not resolved_alpha_path:
            missing.append(alpha_path)
        if not resolved_other_points_path:
            missing.append(other_points_path)
        print(
            "Skipped Sup Table 18/19 because required workbook(s) were not found: "
            + ", ".join(missing)
        )


if __name__ == "__main__":
    generate_workbook()
    print(f"Generated workbook: {OUTPUT_PATH}")
