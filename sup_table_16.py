#!/usr/bin/env python3
"""
Build Supplementary Table 16:
Prevalence of variants with functional impact, normal function, and hypomorphic variants
in secondary structures for BRCA1/BRCA2.

Inputs:
  BRCA1: Sup Table 12 (functional evidence assignment)
  BRCA2: Sup Table 13 (functional evidence assignment)

Notes:
  - T3 (position) and Integrated ACMG evidence criteria are taken from Sup Table 12/13.
  - Variants are only counted as VUS when T6 is empty.

This module is modeled after sup_table_14_15.py, but outputs a single sheet "Sup Table 16"
with two side-by-side gene blocks.
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Iterable, List, Optional

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------------------
# Feature lists
# --------------------------------------------------------------------------------------
try:
    from BRCA1_BRCA2_domains import BRCA1_FEATURES as _BRCA1_FEATURES, BRCA2_FEATURES as _BRCA2_FEATURES
except Exception:  # pragma: no cover - fallback for standalone use
    _BRCA1_FEATURES = []
    _BRCA2_FEATURES = []

# Each entry: {"Feature": "...", "Start aa": int, "End aa": int, "Summary": bool}
BRCA1_FEATURES: List[dict] = list(_BRCA1_FEATURES)
BRCA2_FEATURES: List[dict] = list(_BRCA2_FEATURES)


# --------------------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------------------
BUCKETS = [
    "BS3_supporting",
    "BS3_moderate",
    "BS3",
    "VUS",
    "PS3_supporting",
    "PS3_moderate",
    "PS3",
]

OUTPUT_COLS = [
    "Feature",
    "Start aa",
    "End aa",
    "BS3_supporting",
    "BS3_moderate",
    "BS3",
    "VUS",
    "PS3_supporting",
    "PS3_moderate",
    "PS3",
    "Total",
    "Hypomorph observations",
    "Hypo observations frequency",
    "Frequency of variants with functional impact",
]


def _norm_col(c: str) -> str:
    return re.sub(r"\s+", " ", str(c).strip().lower())


def _find_col(df: pd.DataFrame, candidates: Iterable[str]) -> str:
    mapping = {_norm_col(c): c for c in df.columns}
    for cand in candidates:
        key = _norm_col(cand)
        if key in mapping:
            return mapping[key]
    for cand in candidates:
        key = _norm_col(cand)
        for k_norm, k_orig in mapping.items():
            if key in k_norm:
                return k_orig
    raise KeyError(f"Could not find any of columns: {list(candidates)}")


def resolve_sheet_name(xlsx_path: str, preferred: str, fallbacks: List[str]) -> str:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    names = wb.sheetnames
    if preferred in names:
        return preferred
    lower_map = {n.lower(): n for n in names}
    if preferred.lower() in lower_map:
        return lower_map[preferred.lower()]
    for pat in fallbacks:
        pat_l = pat.lower()
        for n in names:
            if pat_l in n.lower():
                return n
    raise ValueError(f"Sheet not found. preferred={preferred!r}. Available: {names}")


def read_sup_table(xlsx_path: str, sheet_name: str, header_row_index: int = 1) -> pd.DataFrame:
    df = pd.read_excel(xlsx_path, sheet_name=sheet_name, header=header_row_index, engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _normalize_hypo(val: object) -> str:
    return "Y" if str(val).strip().upper() == "Y" else "N"


def _normalize_criteria(val: object) -> str:
    s = str(val).strip()
    if not s or s.lower() in {"nan", "none"}:
        return ""
    s = re.sub(r"[\s\-]+", "_", s.lower())
    s = re.sub(r"[^a-z0-9_]", "", s)
    return s


def bucket_from_criteria(val: object) -> str:
    norm = _normalize_criteria(val)
    mapping = {
        "bs3_supporting": "BS3_supporting",
        "bs3_moderate": "BS3_moderate",
        "bs3": "BS3",
        "ps3_supporting": "PS3_supporting",
        "ps3_moderate": "PS3_moderate",
        "ps3": "PS3",
    }
    return mapping.get(norm, "VUS")


def bucket_from_criteria_strict(val: object) -> Optional[str]:
    norm = _normalize_criteria(val)
    mapping = {
        "bs3_supporting": "BS3_supporting",
        "bs3_moderate": "BS3_moderate",
        "bs3": "BS3",
        "ps3_supporting": "PS3_supporting",
        "ps3_moderate": "PS3_moderate",
        "ps3": "PS3",
    }
    return mapping.get(norm)


def _is_empty(val: object) -> bool:
    if val is None:
        return True
    if isinstance(val, float) and pd.isna(val):
        return True
    return str(val).strip() == ""


def build_vus_universe(variant_df: pd.DataFrame, ref_df: pd.DataFrame) -> pd.DataFrame:
    col_t5 = _find_col(variant_df, ["T5"])
    col_t6 = _find_col(variant_df, ["T6"])
    col_t3 = _find_col(variant_df, ["T3"])
    ref_t5 = _find_col(ref_df, ["T5", "Variant", "Variant ID", "variant", "variant id"])

    ref_variants = set(ref_df[ref_t5].dropna().astype(str).tolist())
    base = variant_df[[col_t5, col_t6, col_t3]].dropna(subset=[col_t5]).copy()
    base[col_t5] = base[col_t5].astype(str)
    vus = base[base[col_t6].isna()].copy()
    vus = vus[~vus[col_t5].isin(ref_variants)]

    out = vus[[col_t5, col_t3]].rename(columns={col_t5: "T5", col_t3: "T3"})
    out["T3"] = pd.to_numeric(out["T3"], errors="coerce")
    return out


def join_functional(vus_df: pd.DataFrame, func_df: pd.DataFrame) -> pd.DataFrame:
    col_t5 = _find_col(func_df, ["T5"])
    col_hypo = _find_col(func_df, ["Hypomorph observation"])
    try:
        col_criteria = _find_col(func_df, ["Integrated ACMG evidence criteria"])
    except KeyError:
        col_criteria = None

    keep_cols = [col_t5, col_hypo] + ([col_criteria] if col_criteria else [])
    func = func_df[keep_cols].dropna(subset=[col_t5]).copy()
    func[col_t5] = func[col_t5].astype(str)

    joined = vus_df.merge(func, how="left", left_on="T5", right_on=col_t5)
    # Drop the right-hand key only when it differs from the left key.
    if col_t5 != "T5" and col_t5 in joined.columns:
        joined = joined.drop(columns=[col_t5])

    joined["Hypomorph observation"] = joined[col_hypo].apply(_normalize_hypo) if col_hypo in joined.columns else "N"
    if col_criteria and col_criteria in joined.columns:
        joined["Integrated ACMG evidence criteria"] = (
            joined[col_criteria].fillna("").astype(str).str.strip()
        )
    else:
        joined["Integrated ACMG evidence criteria"] = ""

    joined["Bucket"] = joined["Integrated ACMG evidence criteria"].map(bucket_from_criteria)
    return joined


def build_assignment_df(sup_df: pd.DataFrame) -> pd.DataFrame:
    col_t5 = _find_col(sup_df, ["T5"])
    col_t3 = _find_col(sup_df, ["T3"])
    col_t6 = _find_col(sup_df, ["T6"])
    col_hypo = _find_col(sup_df, ["Hypomorph observation"])
    col_criteria = _find_col(sup_df, ["Integrated ACMG evidence criteria"])

    df = sup_df[[col_t5, col_t3, col_t6, col_criteria, col_hypo]].dropna(subset=[col_t5]).copy()
    df = df.rename(
        columns={
            col_t5: "T5",
            col_t3: "T3",
            col_t6: "T6",
            col_criteria: "Integrated ACMG evidence criteria",
            col_hypo: "Hypomorph observation",
        }
    )
    df["T5"] = df["T5"].astype(str)
    df["T3"] = pd.to_numeric(df["T3"], errors="coerce")
    df["Integrated ACMG evidence criteria"] = df["Integrated ACMG evidence criteria"].fillna("").astype(str).str.strip()
    df["Hypomorph observation"] = df["Hypomorph observation"].apply(_normalize_hypo)

    buckets: List[Optional[str]] = []
    includes: List[bool] = []
    for _, row in df.iterrows():
        bucket = bucket_from_criteria_strict(row["Integrated ACMG evidence criteria"])
        if bucket:
            buckets.append(bucket)
            includes.append(True)
            continue
        if _is_empty(row["T6"]):
            buckets.append("VUS")
            includes.append(True)
        else:
            buckets.append(None)
            includes.append(False)

    df["Bucket"] = buckets
    df["Include"] = includes
    return df


def _coerce_feature_df(df: pd.DataFrame, gene_label: str) -> pd.DataFrame:
    col_feature = _find_col(df, ["Feature", "Region", "Domain"])
    col_start = _find_col(df, ["Start aa", "Start", "Start_aa"])
    col_end = _find_col(df, ["End aa", "End", "End_aa"])
    summary_col = None
    for cand in ["Summary", "Summary row", "Is summary", "Is_summary"]:
        try:
            summary_col = _find_col(df, [cand])
            break
        except KeyError:
            continue

    out = df[[col_feature, col_start, col_end] + ([summary_col] if summary_col else [])].copy()
    rename_map = {
        col_feature: "Feature",
        col_start: "Start aa",
        col_end: "End aa",
    }
    if summary_col:
        rename_map[summary_col] = "Summary"
    out = out.rename(columns=rename_map)
    if "Summary" not in out.columns:
        out["Summary"] = False
    out["Feature"] = out["Feature"].astype(str).str.strip()
    out["Start aa"] = pd.to_numeric(out["Start aa"], errors="coerce")
    out["End aa"] = pd.to_numeric(out["End aa"], errors="coerce")

    missing = out[out["Start aa"].isna() | out["End aa"].isna()]
    if not missing.empty:
        bad = missing["Feature"].tolist()[:5]
        raise ValueError(
            f"{gene_label} features missing start/end positions. Examples: {bad}"
        )

    out["Summary"] = out["Summary"].apply(
        lambda v: str(v).strip().lower() in {"1", "true", "y", "yes"}
    )
    return out


def load_features(path: Optional[str], default_features: List[dict], gene_label: str) -> pd.DataFrame:
    if path:
        if path.lower().endswith((".xlsx", ".xls")):
            df = pd.read_excel(path)
        else:
            df = pd.read_csv(path)
        return _coerce_feature_df(df, gene_label)

    if not default_features:
        raise ValueError(
            f"No {gene_label} features defined. Populate the {gene_label}_FEATURES list "
            f"or pass --{gene_label.lower()}-features."
        )
    return _coerce_feature_df(pd.DataFrame(default_features), gene_label)


def build_feature_table(assign_df: pd.DataFrame, features_df: pd.DataFrame) -> pd.DataFrame:
    base = assign_df.copy()
    base = base[base["Include"]].dropna(subset=["T3", "T5"])
    rows = []
    for _, feat in features_df.iterrows():
        start = int(feat["Start aa"])
        end = int(feat["End aa"])
        name = feat["Feature"]

        sub = base[(base["T3"] >= start) & (base["T3"] <= end)].copy()
        if not sub.empty:
            sub = sub.drop_duplicates(subset=["T5", "Bucket"])

        counts = sub.groupby("Bucket")["T5"].nunique().to_dict() if len(sub) else {}
        counts = {b: int(counts.get(b, 0)) for b in BUCKETS}

        total = sum(counts.values())
        hypo_count = int(sub[sub["Hypomorph observation"] == "Y"]["T5"].nunique()) if total else 0
        hypo_freq = round((hypo_count / total), 2) if total else 0
        func_freq = round((counts["PS3"] / total), 2) if total else 0

        row = {
            "Feature": name,
            "Start aa": start,
            "End aa": end,
            **counts,
            "Total": total,
            "Hypomorph observations": hypo_count,
            "Hypo observations frequency": hypo_freq,
            "Frequency of variants with functional impact": func_freq,
            "__summary__": bool(feat.get("Summary", False)),
        }
        rows.append(row)

    return pd.DataFrame(rows)


# --------------------------------------------------------------------------------------
# Excel writer
# --------------------------------------------------------------------------------------
def _apply_block_borders(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    thin = Side(style="thin", color="000000")
    thick = Side(style="thick", color="000000")
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            left = thick if c == min_col else thin
            right = thick if c == max_col else thin
            top = thick if r == min_row else thin
            bottom = thick if r == max_row else thin
            ws.cell(row=r, column=c).border = Border(
                left=left, right=right, top=top, bottom=bottom
            )


def _write_block(
    ws,
    start_col: int,
    header_row: int,
    data_row: int,
    df: pd.DataFrame,
    feature_header: str,
    summary_fill: PatternFill,
):
    headers = OUTPUT_COLS.copy()
    headers[0] = feature_header

    header_font = Font(bold=True, name="Calibri", size=10)
    for i, h in enumerate(headers):
        cell = ws.cell(row=header_row, column=start_col + i, value=h)
        cell.font = header_font
        if i == 0:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center", textRotation=90, wrap_text=True)

    for r_idx, row in df.iterrows():
        r = data_row + r_idx
        is_summary = bool(row.get("__summary__", False))
        for c_idx, col in enumerate(OUTPUT_COLS):
            value = row.get(col)
            cell = ws.cell(row=r, column=start_col + c_idx, value=value)
            if c_idx == 0:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            if col in {"Hypo observations frequency", "Frequency of variants with functional impact"}:
                cell.number_format = "0.00"
            elif col not in {"Feature"}:
                cell.number_format = "0"

            if is_summary:
                cell.font = Font(bold=True, name="Calibri", size=10)
                cell.fill = summary_fill
            else:
                cell.font = Font(name="Calibri", size=10)


def write_sup_table_16(
    brca1_df: pd.DataFrame,
    brca2_df: pd.DataFrame,
    output_path: str,
) -> None:
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    if output_file.exists():
        wb = load_workbook(output_file)
        if "Sup Table 16" in wb.sheetnames:
            wb.remove(wb["Sup Table 16"])
        ws = wb.create_sheet("Sup Table 16")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sup Table 16"

    ws.sheet_view.showGridLines = False

    # Layout geometry
    header_row = 2
    data_row = 3
    block_cols = len(OUTPUT_COLS)
    spacer_cols = 1
    left_start = 1
    right_start = left_start + block_cols + spacer_cols
    total_cols = block_cols * 2 + spacer_cols

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title = (
        "Supplementary Table 16: Prevalence of variants with functional impact, "
        "normal function, and hypomorphic variants in secondary structures"
    )
    tcell = ws.cell(row=1, column=1, value=title)
    tcell.font = Font(bold=True, name="Calibri", size=10)
    tcell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    summary_fill = PatternFill("solid", fgColor="E6E6E6")
    _write_block(ws, left_start, header_row, data_row, brca1_df, "BRCA1 Feature", summary_fill)
    _write_block(ws, right_start, header_row, data_row, brca2_df, "BRCA2 Feature", summary_fill)

    # Column widths
    widths = [
        26, 6, 6, 6, 6, 5, 5, 6, 6, 5, 6, 7, 7, 7
    ]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(left_start + i)].width = w
        ws.column_dimensions[get_column_letter(right_start + i)].width = w
    ws.column_dimensions[get_column_letter(left_start + block_cols)].width = 2

    ws.row_dimensions[header_row].height = 120
    ws.freeze_panes = ws["A3"]

    # Borders and separator
    left_max_row = data_row + len(brca1_df) - 1
    right_max_row = data_row + len(brca2_df) - 1
    max_row = max(left_max_row, right_max_row)

    _apply_block_borders(ws, header_row, left_max_row, left_start, left_start + block_cols - 1)
    _apply_block_borders(ws, header_row, right_max_row, right_start, right_start + block_cols - 1)

    # Thicker separator between blocks
    sep_col = left_start + block_cols
    thick = Side(style="thick", color="000000")
    for r in range(header_row, max_row + 1):
        cell = ws.cell(row=r, column=sep_col)
        cell.border = Border(left=thick, right=thick)

    # Conditional formatting for functional impact frequency columns
    scale_rule = ColorScaleRule(
        start_type="min",
        start_color="FFFFFF",
        end_type="max",
        end_color="F4CCCC",
    )
    left_freq_col = left_start + block_cols - 1
    right_freq_col = right_start + block_cols - 1
    ws.conditional_formatting.add(
        f"{get_column_letter(left_freq_col)}{data_row}:{get_column_letter(left_freq_col)}{left_max_row}",
        scale_rule,
    )
    ws.conditional_formatting.add(
        f"{get_column_letter(right_freq_col)}{data_row}:{get_column_letter(right_freq_col)}{right_max_row}",
        scale_rule,
    )

    wb.save(output_file)


# --------------------------------------------------------------------------------------
# CLI
# --------------------------------------------------------------------------------------
def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook", help="Input .xlsx containing Sup Tables 12/13")
    ap.add_argument("-o", "--out", default="SupTable16.xlsx", help="Output .xlsx")
    ap.add_argument("--sup12", default="Sup Table 12")
    ap.add_argument("--sup13", default="Sup Table 13")
    ap.add_argument("--brca1-features", help="CSV/XLSX with BRCA1 features")
    ap.add_argument("--brca2-features", help="CSV/XLSX with BRCA2 features")
    args = ap.parse_args()

    sup12_sheet = resolve_sheet_name(args.workbook, args.sup12, ["sup table 12", "table 12", "functional evidence"])
    sup13_sheet = resolve_sheet_name(args.workbook, args.sup13, ["sup table 13", "table 13", "functional evidence"])

    sup12 = read_sup_table(args.workbook, sup12_sheet, header_row_index=1)
    sup13 = read_sup_table(args.workbook, sup13_sheet, header_row_index=1)

    brca1_features = load_features(args.brca1_features, BRCA1_FEATURES, "BRCA1")
    brca2_features = load_features(args.brca2_features, BRCA2_FEATURES, "BRCA2")

    brca1_assign = build_assignment_df(sup12)
    brca2_assign = build_assignment_df(sup13)

    brca1_tbl = build_feature_table(brca1_assign, brca1_features)
    brca2_tbl = build_feature_table(brca2_assign, brca2_features)

    write_sup_table_16(brca1_tbl, brca2_tbl, args.out)
    print(f"[OK] Wrote Sup Table 16 to {args.out}")


if __name__ == "__main__":
    main()
