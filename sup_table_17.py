#!/usr/bin/env python3
"""
Build Supplementary Table 17:
Fanconi anemia missense variants and likelihood of occurrence.
"""

from __future__ import annotations

import argparse
import math
import re
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

import pandas as pd

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font


FA_VARIANTS = {
    "BRCA1": ["C61G", "C64Y", "R1699W", "R1699Q", "V1736A"],
    "BRCA2": [
        "I3412V", "N372H", "T598A", "E804A", "G106R", "V159M", "R2108C", "R2336H", "R2336P", "I2490T",
        "L2510P", "F2562L", "E2599G", "Y2601C", "S2616F",
        "R2625S", "W2626C", "Q2655R", "S2670L", "A2698T", "D2723H", "D2723V",
        "R2784W", "R2784Q", "G2793R", "R2824T", "R2842C", "E3002K", "G3003E",
        "A3028P", "L3101R",
    ],
}
TOTAL_VARIANTS = {"BRCA1": 3086, "BRCA2": 6100}
FINAL_CODE_CANDIDATES = ("Final functional code label", "Integrated ACMG evidence criteria")
AA3_TO_AA1 = {
    "Ala": "A",
    "Arg": "R",
    "Asn": "N",
    "Asp": "D",
    "Cys": "C",
    "Gln": "Q",
    "Glu": "E",
    "Gly": "G",
    "His": "H",
    "Ile": "I",
    "Leu": "L",
    "Lys": "K",
    "Met": "M",
    "Phe": "F",
    "Pro": "P",
    "Ser": "S",
    "Thr": "T",
    "Trp": "W",
    "Tyr": "Y",
    "Val": "V",
}


def _parse_variant(code: str) -> Tuple[str, int, str]:
    s = str(code).strip()
    m = re.match(r"^(?:p\.)?([A-Za-z])(\d+)([A-Za-z])$", s)
    if not m:
        raise ValueError(f"Unrecognized variant code: {code}")
    return m.group(1).upper(), int(m.group(2)), m.group(3).upper()


def _norm(s: object) -> str:
    return str(s).strip().lower()


def resolve_sheet_name(xlsx_path: str, preferred: str, fallbacks: List[str]) -> str:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    names = wb.sheetnames
    if preferred in names:
        return preferred
    lower_map = {n.lower(): n for n in names}
    if preferred.lower() in lower_map:
        return lower_map[preferred.lower()]
    for pat in fallbacks:
        pat_l = pat.lower()
        for n in names:
            if pat_l in n.lower():
                return n
    raise ValueError(f"Sheet not found. preferred={preferred!r}. Available: {names}")


def _load_sup_table(path: str, sheet: str) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=sheet, header=1)
    df.columns = df.columns.astype(str).str.strip()
    return df


def _dedupe_preserve_order(values: List[str]) -> List[str]:
    seen: set[str] = set()
    out: List[str] = []
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        out.append(value)
    return out


def _load_fa_variants_from_comments(path: str | None) -> Dict[str, List[str]] | None:
    if not path:
        return None
    workbook = Path(path)
    if not workbook.exists():
        return None

    try:
        wb = load_workbook(workbook, read_only=True, data_only=True)
    except Exception:
        return None

    comments_sheet = next((name for name in wb.sheetnames if name.strip().lower() == "comments"), None)
    if comments_sheet is None:
        return None

    ws = wb[comments_sheet]
    variants: Dict[str, List[str]] = {"BRCA1": [], "BRCA2": []}
    in_table_17 = False
    for row in ws.iter_rows(values_only=True):
        first = "" if not row or row[0] is None else str(row[0]).strip()
        first_lower = first.lower()
        if "supplementary table 17" in first_lower:
            in_table_17 = True
            continue
        if in_table_17 and "supplementary table 18" in first_lower:
            break
        if not in_table_17:
            continue
        if len(row) < 6:
            continue

        gene = "" if row[1] is None else str(row[1]).strip().upper()
        wild_type = row[3]
        position = row[4]
        mutant = row[5]
        if gene not in variants or wild_type not in AA3_TO_AA1 or mutant not in AA3_TO_AA1:
            continue

        try:
            pos_int = int(position)
        except (TypeError, ValueError):
            continue

        variants[gene].append(f"{AA3_TO_AA1[wild_type]}{pos_int}{AA3_TO_AA1[mutant]}")

    variants = {gene: _dedupe_preserve_order(items) for gene, items in variants.items()}
    if not all(variants.values()):
        return None
    return variants


def _resolve_fa_variants(path: str | None) -> Dict[str, List[str]]:
    return _load_fa_variants_from_comments(path) or FA_VARIANTS


def _compute_rr_ci(a: int, n1: int, c: int, n0: int) -> Tuple[float, float, float]:
    if n1 == 0 or n0 == 0:
        return 0.0, 0.0, 0.0

    b = n1 - a
    d = n0 - c
    # continuity correction if any zero cell
    if any(x == 0 for x in (a, b, c, d)):
        a += 0.5
        b += 0.5
        c += 0.5
        d += 0.5

    rr = (a / (a + b)) / (c / (c + d)) if c + d else 0.0
    if rr == 0:
        return 0.0, 0.0, 0.0
    se = math.sqrt((1 / a) - (1 / (a + b)) + (1 / c) - (1 / (c + d)))
    lo = math.exp(math.log(rr) - 1.96 * se)
    hi = math.exp(math.log(rr) + 1.96 * se)
    return rr, lo, hi


def _category_masks(df: pd.DataFrame) -> Dict[str, pd.Series]:
    conc = df["Concordance"].astype(str).str.strip().str.lower()
    prepon = df["Preponderance of evidence"].astype(str).str.strip().str.lower()
    criteria_col = next((c for c in FINAL_CODE_CANDIDATES if c in df.columns), None)
    if criteria_col is None:
        raise KeyError(f"Missing final functional code column. Expected one of: {FINAL_CODE_CANDIDATES}")
    criteria = df[criteria_col].astype(str).str.strip().str.lower()
    hypo = df["Hypomorph observation"].astype(str).str.strip().str.upper()

    return {
        "Unresolved discordant set": (conc == "discordant") & (prepon == "indeterminate"),
        "Indeterminate": criteria == "indeterminate",
        "Functional Impact set (PS3_supporting, PS3_moderate, PS3)": criteria.isin(
            {"ps3_supporting", "ps3_moderate", "ps3"}
        ),
        "Normal Function set (BS3_supporting, BS3_moderate, BS3)": criteria.isin(
            {"bs3_supporting", "bs3_moderate", "bs3"}
        ),
        "Hypomorph observation": hypo == "Y",
    }


def _count_variants(df: pd.DataFrame, keys: set[Tuple[str, int, str]], mask: pd.Series) -> int:
    sub = df[mask].dropna(subset=["T2", "T3", "T4"]).copy()
    sub["T2"] = sub["T2"].astype(str).str.strip().str.upper()
    sub["T4"] = sub["T4"].astype(str).str.strip().str.upper()
    sub["T3"] = pd.to_numeric(sub["T3"], errors="coerce")
    sub = sub.dropna(subset=["T3"])
    sub = sub.drop_duplicates(subset=["T2", "T3", "T4"])
    if not keys:
        return len(sub)
    in_keys = sub.apply(lambda r: (r["T2"], int(r["T3"]), r["T4"]) in keys, axis=1)
    return int(in_keys.sum())


def _write_block(ws, start_row: int, gene_label: str, fa_n: int, all_n: int) -> int:
    # Header row for the gene block
    ws.cell(row=start_row, column=1, value=gene_label).font = Font(bold=True)
    for col, value in [
        (2, f"Fraction of FA variants (n = {fa_n})"),
        (3, f"Fraction of all variants (n = {all_n})"),
        (4, "LR"),
        (5, "95% CI"),
    ]:
        cell = ws.cell(row=start_row, column=col, value=value)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    rows: List[str] = [
        "Unresolved discordant set",
        "Indeterminate",
        "Functional Impact set (PS3_supporting, PS3_moderate, PS3)",
        "Normal Function set (BS3_supporting, BS3_moderate, BS3)",
        "Hypomorph observation",
    ]

    r = start_row + 1
    for label in rows:
        cell = ws.cell(row=r, column=1, value=label)
        cell.font = Font(bold=True)
        r += 1
    return r


def build_table(
    sup_df: pd.DataFrame, gene: str, fa_variants: List[str], fa_total: int, all_total: int
) -> List[Tuple[str, float, float, float, str]]:
    keys = {_parse_variant(v) for v in fa_variants}
    masks = _category_masks(sup_df)

    rows = []
    for label, mask in masks.items():
        fa_count = _count_variants(sup_df, keys, mask)
        all_count = _count_variants(sup_df, set(), mask)
        fa_frac = fa_count / fa_total if fa_total else 0.0
        all_frac = all_count / all_total if all_total else 0.0
        lr, ci_lo, ci_hi = _compute_rr_ci(fa_count, fa_total, all_count, all_total)
        rows.append(
            (
                label,
                round(fa_frac, 2),
                round(all_frac, 2),
                round(lr, 2),
                f"{ci_lo:.2f} to {ci_hi:.2f}",
            )
        )
    return rows


def write_sup_table_17(
    workbook: str,
    output_path: str,
    sup12_sheet: str = "Sup Table 12",
    sup13_sheet: str = "Sup Table 13",
    comments_workbook: str | None = None,
) -> None:
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    sup12_name = resolve_sheet_name(workbook, sup12_sheet, ["sup table 12", "table 12"])
    sup13_name = resolve_sheet_name(workbook, sup13_sheet, ["sup table 13", "table 13"])
    sup12 = _load_sup_table(workbook, sup12_name)
    sup13 = _load_sup_table(workbook, sup13_name)
    fa_variants = _resolve_fa_variants(comments_workbook or workbook)
    fa_totals = {gene: len(variants) for gene, variants in fa_variants.items()}

    if output_file.exists():
        wb = load_workbook(output_file)
        for sheet_name in ("Sup Table 17", "Supp Table 17"):
            if sheet_name in wb.sheetnames:
                wb.remove(wb[sheet_name])
        ws = wb.create_sheet("Supp Table 17")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Supp Table 17"

    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:E1")
    ws["A1"] = "Supplementary Table 17: Fanconi anemia missense variants and likelihood of occurrence"
    ws["A1"].font = Font(bold=True)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    brca1_rows = build_table(
        sup12, "BRCA1", fa_variants["BRCA1"], fa_totals["BRCA1"], TOTAL_VARIANTS["BRCA1"]
    )
    brca2_rows = build_table(
        sup13, "BRCA2", fa_variants["BRCA2"], fa_totals["BRCA2"], TOTAL_VARIANTS["BRCA2"]
    )

    row = 3
    row = _write_block(ws, row, "BRCA1", fa_totals["BRCA1"], TOTAL_VARIANTS["BRCA1"])
    for i, (label, fa_frac, all_frac, lr, ci) in enumerate(brca1_rows, start=0):
        r = 4 + i
        ws.cell(row=r, column=2, value=fa_frac)
        ws.cell(row=r, column=3, value=all_frac)
        ws.cell(row=r, column=4, value=lr)
        ws.cell(row=r, column=5, value=ci)

    row = 11
    _write_block(ws, row, "BRCA2", fa_totals["BRCA2"], TOTAL_VARIANTS["BRCA2"])
    for i, (label, fa_frac, all_frac, lr, ci) in enumerate(brca2_rows, start=0):
        r = 12 + i
        ws.cell(row=r, column=2, value=fa_frac)
        ws.cell(row=r, column=3, value=all_frac)
        ws.cell(row=r, column=4, value=lr)
        ws.cell(row=r, column=5, value=ci)

    # Basic alignment/widths
    for col, width in zip(["A", "B", "C", "D", "E"], [60, 18, 18, 10, 18]):
        ws.column_dimensions[col].width = width
    for r in ws.iter_rows():
        for c in r:
            if c.value is None:
                continue
            if c.column in (2, 3, 4, 5):
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")
            if c.column in (2, 3, 4):
                c.number_format = "0.00"

    wb.save(output_file)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook", help="Input workbook containing Sup Table 12/13")
    ap.add_argument("-o", "--out", default="SupTable17.xlsx", help="Output .xlsx")
    ap.add_argument("--sup12", default="Sup Table 12")
    ap.add_argument("--sup13", default="Sup Table 13")
    ap.add_argument("--comments-workbook", help="Optional workbook containing the Comments tab with the FA variant list")
    args = ap.parse_args()
    write_sup_table_17(
        args.workbook,
        args.out,
        sup12_sheet=args.sup12,
        sup13_sheet=args.sup13,
        comments_workbook=args.comments_workbook,
    )
    print(f"[OK] Wrote Sup Table 17 to {args.out}")


if __name__ == "__main__":
    main()
